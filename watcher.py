import os
import time
import logging
import threading
import requests
import re
from io import BytesIO
from openpyxl import load_workbook

# ================= CONFIG =================

MS_TENANT_ID = os.getenv("TENANT_ID")
MS_CLIENT_ID = os.getenv("CLIENT_ID_DRIVE")
MS_CLIENT_SECRET = os.getenv("CLIENT_SECRET_DRIVE")

DRIVE_ID = os.getenv("DRIVE_ID")
ITEM_ID = os.getenv("ITEM_ID")

POLL_INTERVAL = int(os.getenv("POLL_INTERVAL", 60))

GRAPH_TOKEN_URL = f"https://login.microsoftonline.com/{MS_TENANT_ID}/oauth2/v2.0/token"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)

ID_RE = re.compile(r"_(\d{4})$")
# ================= TOKEN =================

class TokenManager:
    def __init__(self):
        self.token = None
        self.expiry = 0
        self.lock = threading.Lock()

    def get_token(self):
        with self.lock:

            if self.token and time.time() < self.expiry - 300:
                return self.token

            logging.info("Refreshing Microsoft token...")

            resp = requests.post(
                GRAPH_TOKEN_URL,
                data={
                    "client_id": MS_CLIENT_ID,
                    "client_secret": MS_CLIENT_SECRET,
                    "grant_type": "client_credentials",
                    "scope": "https://graph.microsoft.com/.default",
                },
                timeout=30,
            )

            resp.raise_for_status()
            data = resp.json()

            self.token = data["access_token"]
            self.expiry = time.time() + int(data["expires_in"])

            return self.token


token_manager = TokenManager()


def graph_headers():
    return {"Authorization": f"Bearer {token_manager.get_token()}"}


# ================= RETRY =================

def request_with_retry(method, url, **kwargs):
    for attempt in range(5):
        try:
            resp = requests.request(method, url, timeout=30, **kwargs)

            if resp.status_code == 429:
                wait = int(resp.headers.get("Retry-After", 5))
                logging.warning(f"Throttled. Sleeping {wait}s")
                time.sleep(wait)
                continue

            if resp.status_code >= 500:
                raise requests.RequestException(resp.text)

            resp.raise_for_status()
            return resp

        except Exception as e:
            wait = 2 ** attempt
            logging.warning(f"Request failed: {e}. Retry in {wait}s")
            time.sleep(wait)

    raise Exception("Max retries exceeded")


# ================= GRAPH =================

def get_last_modified():
    url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{ITEM_ID}"

    resp = request_with_retry("GET", url, headers=graph_headers())

    return resp.json()["lastModifiedDateTime"]


def download_excel():
    url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{ITEM_ID}/content"

    resp = request_with_retry("GET", url, headers=graph_headers())

    return resp.content
def get_file_metadata():
    url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{ITEM_ID}"

    resp = request_with_retry("GET", url, headers=graph_headers())
    data = resp.json()

    return data["lastModifiedDateTime"], data["eTag"]

def upload_excel(data, etag):
    url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{ITEM_ID}/content"

    headers = graph_headers()
    headers["If-Match"] = etag   # 🔥 THE MAGIC LINE

    for attempt in range(10):  # try for up to ~2 minutes
        resp = requests.put(url, headers=headers, data=data, timeout=30)

        if resp.status_code == 423:
            logging.info("File locked during upload. Waiting 10s...")
            time.sleep(10)
            continue

        if resp.status_code == 412:
            logging.warning("Upload rejected — file changed during processing.")
            return False

        resp.raise_for_status()
        return True

    logging.warning("Upload skipped — file remained locked.")
    return False


# ================= EXCEL =================

def assign_ids(file_bytes):
    wb = load_workbook(BytesIO(file_bytes))
    global_ws = wb["__GLOBAL__"]

    TARGET_COLUMNS = [1, 9]  # B and J
    used_ids = set()

    # -------- FIRST PASS: Collect existing IDs --------
    for ws in wb.worksheets:
        if ws.title == "__GLOBAL__":
            continue

        for row in ws.iter_rows():
            for col_index in TARGET_COLUMNS:
                if col_index < len(row):
                    value = row[col_index].value
                    if isinstance(value, str):
                        m = ID_RE.search(value)
                        if m:
                            used_ids.add(int(m.group(1)))

    last_id = max(used_ids) if used_ids else 0
    changed = False

    # -------- SECOND PASS: Assign new IDs --------
    for ws in wb.worksheets:
        if ws.title == "__GLOBAL__":
            continue

        for row in ws.iter_rows():
            for col_index in TARGET_COLUMNS:

                if col_index >= len(row):
                    continue

                cell = row[col_index]
                name = cell.value

                if not isinstance(name, str) or not name.strip():
                    continue

                if ID_RE.search(name):
                    continue

                if name.strip().lower() in {
                    "მოსალოდნელი პროექტი",
                    "მიმდინარე პროექტი",
                    "პროექტი"
                }:
                    continue

                # Generate next free ID
                new_id = max(used_ids, default=0) + 1

                used_ids.add(new_id)
                cell.value = f"{name}_{new_id:04d}"
                changed = True
                last_id = max(last_id, new_id)

    # -------- SAVE IF CHANGED --------
    if changed:
        global_ws["B1"].value = last_id

        out = BytesIO()
        wb.save(out)
        out.seek(0)

        return out.read(), last_id

    return None, last_id


# ================= WATCHER =================

def watcher_loop():
    logging.info("Watcher started.")

    STABILITY_SECONDS = 20  # Wait this long without changes
    last_seen = None
    last_change_time = None

    try:
        last_seen, _ = get_file_metadata()
    except Exception:
        logging.exception("Initial metadata fetch failed.")

    while True:
        try:
            modified, etag = get_file_metadata()

            # File changed
            if modified != last_seen:
                logging.info("Change detected. Waiting for stability...")
                last_seen = modified
                last_change_time = time.time()

            # If change was detected earlier, wait until stable
            if last_change_time:
                time_since_change = time.time() - last_change_time

                # Check if file changed again during waiting
                current_modified, _ = get_file_metadata()

                if current_modified != last_seen:
                    # File changed again — reset timer
                    last_seen = current_modified
                    last_change_time = time.time()
                    logging.info("File changed again. Resetting stability timer.")
                
                elif time_since_change >= STABILITY_SECONDS:
                    # File is stable — now process
                    logging.info("File stable. Processing...")

                    file_bytes = download_excel()
                    result, last_id = assign_ids(file_bytes)

                    if result:
                        _, fresh_etag = get_file_metadata()
                        success = upload_excel(result, fresh_etag)
                        if success:
                            logging.info(f"IDs assigned. Last ID = {last_id}")
                            last_seen = modified
                        else:
                            logging.info("Upload failed (likely locked). Will retry next cycle.")
                    if success:
                        last_change_time = None  # reset timer

        except Exception:
            logging.exception("Watcher loop error")

        time.sleep(POLL_INTERVAL)
_watcher_thread = None

def start_watcher():
    global _watcher_thread

    if _watcher_thread and _watcher_thread.is_alive():
        logging.info("Watcher already running.")
        return

    _watcher_thread = threading.Thread(
        target=watcher_loop,
        daemon=True
    )
    _watcher_thread.start()

    logging.info("Background watcher started.")