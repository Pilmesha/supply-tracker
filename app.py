import os, requests, hmac, hashlib, io, random, time, threading, gc, base64, re, pdfplumber
from flask import Flask, request, jsonify, make_response
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from datetime import datetime, timedelta, date
from concurrent.futures import ThreadPoolExecutor, as_completed
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import traceback
from collections import defaultdict
import sqlite3
from watcher import start_watcher
load_dotenv()

#======CONGIF=====

# single session (reuse connections)
HTTP = requests.Session()
HTTP.headers.update({"User-Agent": "supply-tracker/1.0", "Content-Type": "application/x-www-form-urlencoded"})
retry_strategy = Retry(
    status_forcelist=[429, 500, 502, 503, 504],
    allowed_methods=["HEAD", "GET", "OPTIONS", "POST", "PUT", "PATCH", "DELETE"]
)
adapter = HTTPAdapter(max_retries=retry_strategy, pool_connections=10, pool_maxsize=10)
HTTP.mount("https://", adapter)
HTTP.mount("http://", adapter)
# thread pool to avoid unbounded thread creation
POOL = ThreadPoolExecutor(max_workers=4)  # tune 2-4 on free tier
# single lock to avoid concurrent workbook uploads
EXCEL_LOCK = threading.Lock()

CLIENT_ID = os.getenv('CLIENT_ID')
CLIENT_SECRET = os.getenv('CLIENT_SECRET')
REFRESH_TOKEN = os.getenv('REFRESH_TOKEN')
ORG_ID = os.getenv('ORG_ID')
TENANT_ID = os.getenv('TENANT_ID')
CLIENT_ID_DRIVE = os.getenv('CLIENT_ID_DRIVE')
CLIENT_SECRET_DRIVE = os.getenv('CLIENT_SECRET_DRIVE')
DRIVE_ID = os.getenv('DRIVE_ID')
FILE_ID = os.getenv('FILE_ID')
PERMS_ID = os.getenv('PERMS_ID')
HACH_FILE = os.getenv('HACH_FILE')
HACH_HS = os.getenv('HACH_HS')
TRANS_FILE = os.getenv("TRANS_FILE")
ACCESS_TOKEN_DRIVE = None
ACCESS_TOKEN_EXPIRY = datetime.utcnow()
ACCESS_TOKEN = None

MAILBOXES = [
    "info@vortex.ge",
    "archil@vortex.ge",
    "Logistics@vortex.ge",
    "hach@vortex.ge"
]
MAILBOXES_2 = [
    "info@vortex.ge",
    "teona@vortex.ge"
]
WEBHOOK_URL = "https://supply-tracker-o7ro.onrender.com/webhook"
GRAPH_URL = "https://graph.microsoft.com/v1.0"
DB_PATH = os.environ.get("SQLITE_DB_PATH", "/tmp/processed.db")
conn = sqlite3.connect(DB_PATH, check_same_thread=False)
cursor = conn.cursor()

cursor.execute("""
CREATE TABLE IF NOT EXISTS processed_messages (
    internet_id TEXT PRIMARY KEY
)
""")
conn.commit()


app = Flask(__name__)
start_watcher()

# ======= AUTH ===========
def refresh_access_token() -> str:
    global ACCESS_TOKEN

    url = "https://accounts.zoho.com/oauth/v2/token"
    params = {
        "refresh_token": REFRESH_TOKEN,
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "grant_type": "refresh_token"
    }

    resp = HTTP.post(url, params=params)
    resp.raise_for_status()

    data = resp.json()

    if "access_token" not in data:
        raise Exception(f"Zoho token refresh failed: {data}")

    ACCESS_TOKEN = data["access_token"]
    return ACCESS_TOKEN
def verify_zoho_signature(request, expected_module):
    # Select secret based on webhook type
    secret_key = (
    os.getenv("PURCHASE_WEBHOOK_SECRET")
    if expected_module == "purchaseorders"
    else os.getenv("RECEIVE_WEBHOOK_SECRET")
    if expected_module == "purchasereceive"
    else os.getenv("INVOICE_WEBHOOK_SECRET")
    if expected_module == "invoice"
    else os.getenv("SHIPMENT_WEBHOOK_SECRET")
    
    ).encode("utf-8")
    
    received_sign = request.headers.get('X-Zoho-Webhook-Signature')
    if not received_sign or not secret_key:
        return False
    
    expected_sign = hmac.new(
        secret_key,
        request.get_data(),
        hashlib.sha256
    ).hexdigest()
    
    return hmac.compare_digest(received_sign, expected_sign)
def One_Drive_Auth() -> str:
    global ACCESS_TOKEN_DRIVE, ACCESS_TOKEN_EXPIRY
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    data = {
        "grant_type": "client_credentials",
        "client_id": CLIENT_ID_DRIVE,
        "client_secret": CLIENT_SECRET_DRIVE,
        "scope": "https://graph.microsoft.com/.default"
    }
    try:
        resp = HTTP.post(url, data=data, timeout=30)
        resp.raise_for_status()
        response_json = resp.json()
        
        ACCESS_TOKEN_DRIVE = response_json.get("access_token")
        expires_in = response_json.get("expires_in", 3600)
        ACCESS_TOKEN_EXPIRY = datetime.utcnow() + timedelta(seconds=expires_in - 60)  # refresh 1 min early
        
        if ACCESS_TOKEN_DRIVE:
            return ACCESS_TOKEN_DRIVE
        else:
            print("No access_token in response!")
            return None
    except Exception as e:
        print(f"Error getting access token: {e}")
        return None
def get_headers():
    global ACCESS_TOKEN_DRIVE, ACCESS_TOKEN_EXPIRY
    if (ACCESS_TOKEN_DRIVE is None) or (ACCESS_TOKEN_EXPIRY <= datetime.utcnow()):
        One_Drive_Auth()  # refresh token + expiry
    return {
        "Authorization": f"Bearer {ACCESS_TOKEN_DRIVE}",
        "Content-Type": "application/json"
    }

# =========== HELPER FUNCS FOR EXCEL =============
def get_used_range(sheet_name: str):
    url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{FILE_ID}/workbook/worksheets/{sheet_name}/usedRange"
    headers = {"Authorization": f"Bearer {ACCESS_TOKEN_DRIVE}"}
    resp = HTTP.get(url, headers=headers, params={"valuesOnly": "false"})
    resp.raise_for_status()
    return resp.json()["address"]  # e.g. "მიმდინარე !A1:Y20"
def create_table_if_not_exists(range_address, sheet_name, has_headers=True, retries=3):
    headers = {"Authorization": f"Bearer {ACCESS_TOKEN_DRIVE}"}

    # ✅ 1. Query ONLY tables from the specified sheet
    url_sheet_tables = (
        f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{FILE_ID}"
        f"/workbook/worksheets/{sheet_name}/tables"
    )

    resp = HTTP.get(url_sheet_tables, headers=headers)
    resp.raise_for_status()
    sheet_tables = resp.json().get("value", [])

    # If any table exists on sheet → reuse first table
    if sheet_tables:
        return sheet_tables[0]["name"]

    # --- Create a new table ---
    url_add = (
        f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{FILE_ID}/workbook/tables/add"
    )
    headers["Content-Type"] = "application/json"
    payload = {"address": range_address, "hasHeaders": has_headers}

    for _ in range(retries):
        resp = HTTP.post(url_add, headers=headers, json=payload)
        if resp.status_code in [200, 201]:
            table = resp.json()
            print(f"✅ Created table '{table['name']}' at {range_address}")
            return table["name"]
        else:
            print(f"⚠️ Table creation failed ({resp.status_code}), retrying...")
            time.sleep(2)

    raise Exception(
        f"❌ Failed to create table after {retries} retries: "
        f"{resp.status_code} {resp.text}"
    )
def get_table_columns(table_name):
    url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{FILE_ID}/workbook/tables/{table_name}/columns"
    headers = {"Authorization": f"Bearer {ACCESS_TOKEN_DRIVE}"}
    resp = HTTP.get(url, headers=headers)
    resp.raise_for_status()
    return [col["name"] for col in resp.json().get("value", [])]
def delete_table_rows(sheet_name: str, row_numbers: list[int]):
    headers = {
        "Authorization": f"Bearer {ACCESS_TOKEN_DRIVE}",
        "Content-Type": "application/json"
    }

    for row in sorted(row_numbers, reverse=True):
        address = f"{row}:{row}"  # delete whole row
        url = (
            f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{FILE_ID}"
            f"/workbook/worksheets/{sheet_name}/range(address='{address}')/delete"
        )

        resp = HTTP.post(url, headers=headers, json={"shift": "up"})
        if resp.status_code not in (200, 204):
            print(f"⚠️ Failed to delete row {row}: {resp.text}")
        else:
            print(f"🗑️ Deleted worksheet row {row}")
def get_table_start_row_from_used_range(sheet_name: str) -> int:
    used_addr = get_used_range(sheet_name)
    # Example: "მიმდინარე !A1:Y300"
    start_cell = used_addr.split("!")[1].split(":")[0]  # "A1"
    start_row = int(re.findall(r"\d+", start_cell)[0])
    return start_row
def normalize_hach(df: pd.DataFrame) -> pd.DataFrame:
    table_cols = [
        "Item", "წერილი", "Code", "HS Code", "Details", "თარგმანი", "QTY",
        "მიწოდების ვადა", "Confirmation 1 (shipment week)", "Packing List",
        "რა რიცხვში გამოგზავნეს Packing List-ი", "რამდენი გამოიგზავნა",
        "ჩამოსვლის სავარაუდო თარიღი", "რეალური ჩამოსვლის თარიღი",
        "Qty Delivered", "Customer", "Export?", "მდებარეობა", "შენიშვნა"
    ]

    # --- Base shaping ---
    df = df[['Item', 'Code', 'შეკვეთილი რაოდენობა', 'Customer', 'Export?', 'მიწოდების ვადა']].copy()
    df = df.rename(columns={"Item": "Details", "შეკვეთილი რაოდენობა": "QTY"})
    df["Item"] = df.index + 1

    # --- Download reference files ---
    headers = {"Authorization": f"Bearer {ACCESS_TOKEN_DRIVE or One_Drive_Auth()}"}

    def download_excel(file_id):
        url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{file_id}/content"
        resp = HTTP.get(url, headers=headers, timeout=60)
        resp.raise_for_status()
        return io.BytesIO(resp.content)

    hs_stream     = download_excel(HACH_HS)
    letter_stream = download_excel(PERMS_ID)

    # --- HS codes ---
    hs_df = pd.read_excel(hs_stream, header=[0,1])
    hs_work = hs_df.loc[:, [1, 16, 19, 25]].copy()

    # Rename internally to match your existing logic
    hs_work.columns = ["Code", "GL ID", "ID", "HS Code"]
    hs_work["Code"] = hs_work["Code"].astype(str).str.strip()
    hs_work["HS Code"] = (hs_work["HS Code"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True))
    # --- Permissions ---
    letter_df = pd.read_excel(letter_stream, header=1)
    letter_stream.close()
    # --- Translations ---
    url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{TRANS_FILE}/content"
    resp = HTTP.get(url, headers=headers, timeout=60)
    resp.raise_for_status()
    trans = pd.read_excel(io.BytesIO(resp.content))
    trans_lookup = {}
    for _, row in trans.iterrows():
        if pd.notna(row['Item']) and pd.notna(row['თარგმანი']):
            # Normalize item by removing punctuation
            normalized_item = str(row['Item']).translate(str.maketrans('', '', '.,\n\r\t')).lower().strip()
            trans_lookup[normalized_item] = row['თარგმანი']
    def get_translation(item):
        if pd.isna(item):
            return ""
        # Normalize the item text by removing punctuation
        normalized = str(item).translate(str.maketrans('', '', '.,\n\r\t')).lower().strip()
        return trans_lookup.get(normalized, "")

    if {"მწარმოებლის კოდი", "მიღებული ნებართვა 1 / წერილის ნომერი"}.issubset(letter_df.columns):
        letter_df = letter_df[
            ["მწარმოებლის კოდი", "მიღებული ნებართვა 1 / წერილის ნომერი"]
        ]
        letter_df["მწარმოებლის კოდი"] = (
            letter_df["მწარმოებლის კოდი"].astype(str).str.strip()
        )
    else:
        letter_df = pd.DataFrame(columns=["მწარმოებლის კოდი", "მიღებული ნებართვა 1 / წერილის ნომერი"])

    # --- Ensure all target columns exist ---
    for col in table_cols:
        if col not in df.columns:
            df[col] = ""

    # --- Normalize Code once ---
    df["Code"] = df["Code"].astype(str).str.strip()

    # --- Build lookups ---
    hs_lookup = (
        hs_work
        .drop_duplicates(subset="Code")
        .set_index("Code")["HS Code"]
    )

    perm_lookup = (
        letter_df
        .drop_duplicates(subset="მწარმოებლის კოდი")
        .set_index("მწარმოებლის კოდი")["მიღებული ნებართვა 1 / წერილის ნომერი"]
    )
    is_reag = hs_work[(hs_work['ID'] == "Chemistry") & (hs_work["GL ID"] == "Chemical")]
    reag_codes = set(is_reag["Code"].astype(str).str.strip())
    # --- Fill EXISTING columns only ---
    df["HS Code"] = df["Code"].map(hs_lookup)
    df["წერილი"] = df["Code"].map(perm_lookup)
    mask_reag_no_letter = (
        df["Code"].isin(reag_codes) &
        df["წერილი"].isna()
    )
    df.loc[mask_reag_no_letter, "წერილი"] = "შესატანია"
    # non-reagent + no letter → "არ სჭირდება"
    df["წერილი"] = df["წერილი"].fillna("არ სჭირდება")
    # --- Final column order ---
    # --- Fill translations for Details column ---
    df["თარგმანი"] = df["Details"].apply(get_translation)
    df['მდებარეობა'] = "გერმანია"
    df = df[table_cols]
    return df.fillna("").astype(str)
def split_pdf_by_po(pdf_text: str, po_numbers: list[str]) -> dict[str, str]:
    blocks = {}
    # sort by PO occurrence in PDF
    po_positions = []

    # Find start position of each PO in PDF
    for po in po_numbers:
        # regex to find PO with optional leading zeros
        match = re.search(rf"PO\s*[-:#–]?\s*0*{po}\b", pdf_text)
        if match:
            po_positions.append((po, match.start()))
        else:
            print(f"⚠️ PO-{po} not found in PDF text")
    
    # sort by start index
    po_positions.sort(key=lambda x: x[1])

    for i, (po, start) in enumerate(po_positions):
        if i + 1 < len(po_positions):
            end = po_positions[i + 1][1]
        else:
            end = len(pdf_text)
        blocks[po] = pdf_text[start:end]

    return blocks
def graph_safe_request(method, url, headers, json=None, max_retries=5):
    last_resp = None

    for attempt in range(max_retries):
        try:
            resp = safe_request(
                method,
                url,
                headers=headers,
                json=json,
                timeout=30
            )
            last_resp = resp
            status = resp.status_code

            if status < 400:
                return resp

            if status not in (423, 429) and status < 500:
                resp.raise_for_status()

            print(
                f"⚠️ Graph busy (HTTP {status}), retry {attempt + 1}/{max_retries}"
            )
            time.sleep(1 + attempt * 1.5)

        except requests.Timeout:
            print(
                f"⏱️ Graph timeout, retry {attempt + 1}/{max_retries}"
            )
            time.sleep(1 + attempt * 1.5)

        except requests.RequestException as e:
            print(
                f"⚠️ Graph exception: {e}, retry {attempt + 1}/{max_retries}"
            )
            time.sleep(1 + attempt * 1.5)

    print(f"❌ Graph failed after {max_retries} retries")

    if last_resp is not None:
        last_resp.raise_for_status()
    else:
        raise RuntimeError("Graph request failed with no response returned.")
def is_empty(val):
    return val is None or (isinstance(val, float) and pd.isna(val)) or str(val).strip() == ""
def extract_po_k_mapping(pdf_text: str) -> dict:
    po_pattern = re.compile(r"\bPO[-:#]?\s*(\d+)\b")
    k_pattern = re.compile(r"\bK\d{9}\b", re.IGNORECASE)

    po_matches = list(po_pattern.finditer(pdf_text))
    mapping = {}

    if not po_matches:
        print("❌ No PO numbers found in text")
        return mapping

    # Find all Ks in the document
    all_k = k_pattern.findall(pdf_text)
    first_k = all_k[0].upper() if all_k else None

    for idx, po in enumerate(po_matches):
        po_digits = str(int(po.group(1)))
        block_start = po.end()

        # block ends at next PO or end of document
        block_end = po_matches[idx + 1].start() if idx + 1 < len(po_matches) else len(pdf_text)
        po_block = pdf_text[block_start:block_end]

        # Try to find K inside the PO block
        k_match = k_pattern.search(po_block)
        if k_match:
            mapping[po_digits] = k_match.group(0).upper()
        elif first_k:
            # fallback to first K in the document
            mapping[po_digits] = first_k
            print(f"⚠️ No K found inside PO-{po_digits} block, using first K in document")
        else:
            print(f"⚠️ No K found for PO-{po_digits} anywhere")

    return mapping
def get_sheet_values(sheet_name: str):
    url = (
        f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/"
        f"{FILE_ID}/workbook/worksheets/{sheet_name}/usedRange?$select=values"
    )
    
    headers = {"Authorization": f"Bearer {ACCESS_TOKEN_DRIVE}"}
    
    resp = HTTP.get(url, headers=headers)
    resp.raise_for_status()

    result = resp.json()
    return result.get("values", [])  # this is the list of rows
def format_hach_sheet_full(sheet_name: str,start_row: int,row_count: int,table_id: str) -> None:
    headers = {
        "Authorization": f"Bearer {ACCESS_TOKEN_DRIVE}",
        "Content-Type": "application/json"
    }

    last_row = start_row + row_count
    info_range  = "C3:D6"
    data_range  = f"B{start_row + 1}:T{last_row}"
    header_range = f"B{start_row}:T{start_row}"


    base_url = (
        f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}"
        f"/items/{HACH_FILE}/workbook"
    )
    # ---------------------------- # 1. INFO BLOCK (C3:D6) # ---------------------------- 
    # Alignment
    graph_safe_request(
        "PATCH",
        f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{HACH_FILE}"
        f"/workbook/worksheets/{sheet_name}/range(address='{info_range}')/format",
        headers,
        {"verticalAlignment": "Center", "horizontalAlignment": "Center"}
    ).raise_for_status()

    # Borders
    for edge in [
        "EdgeTop", "EdgeBottom", "EdgeLeft",
        "EdgeRight", "InsideHorizontal", "InsideVertical"
    ]:
        graph_safe_request(
            "PATCH",
            f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{HACH_FILE}"
            f"/workbook/worksheets/{sheet_name}/range(address='{info_range}')"
            f"/format/borders/{edge}",
            headers,
            {"style": "Continuous", "weight": "Thin", "color": "#000000"}
        ).raise_for_status()
    # -------------------------------------------------
    # HEADER FORMAT
    # -------------------------------------------------
    graph_safe_request(
        "PATCH",
        f"{base_url}/worksheets/{sheet_name}"
        f"/range(address='{header_range}')/format",
        headers,
        {
            "horizontalAlignment": "Center",
            "verticalAlignment": "Center",
            "wrapText": True
        }
    ).raise_for_status()
    graph_safe_request(
    "PATCH",
    f"{base_url}/worksheets/{sheet_name}"
    f"/range(address='{sheet_name}!{start_row}:{start_row}')/format",
    headers,
    {"rowHeight": 20}
    ).raise_for_status()


    # -------------------------------------------------
    # 2. Format ALL data cells at once
    # -------------------------------------------------
    graph_safe_request(
        "PATCH",
        f"{base_url}/worksheets/{sheet_name}"
        f"/range(address='{data_range}')/format",
        headers,
        {
            "horizontalAlignment": "Center",
            "verticalAlignment": "Center",
            "wrapText": True
        }
    ).raise_for_status()

    # -------------------------------------------------
    # 3. Set row height ONCE
    # -------------------------------------------------
    graph_safe_request(
        "PATCH",
        f"{base_url}/worksheets/{sheet_name}"
        f"/range(address='{sheet_name}!{start_row + 1}:{last_row}')/format",
        headers,
        {"rowHeight": 35}
    ).raise_for_status()

    # -------------------------------------------------
    # 4. Set column widths in ONE call (grouped)
    # -------------------------------------------------
    width_map = {
        "B": 45, "C": 120, "D": 110, "E": 110,
        "F": 300, "G": 140, "H": 60, "I": 120,
        "J": 160, "K": 160, "L": 180, "M": 160,
        "N": 160, "O": 120, "P": 120, "Q": 80,
        "R": 120, "S": 140, "T": 200
    }

    for col, width in width_map.items():
        graph_safe_request(
            "PATCH",
            f"{base_url}/worksheets/{sheet_name}"
            f"/range(address='{sheet_name}!{col}:{col}')/format",
            headers,
            {"columnWidth": width}
        ).raise_for_status()
    print("🎨 HACH formatting applied")
def load_hach_reference_values() -> set[str]:
    url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{HACH_HS}/content"
    headers = {
        "Authorization": f"Bearer {ACCESS_TOKEN_DRIVE or One_Drive_Auth()}"
    }

    resp = HTTP.get(url, headers=headers, timeout=60)
    resp.raise_for_status()

    wb = load_workbook(io.BytesIO(resp.content), read_only=True)
    ws = wb.active  # assume first sheet

    hach_values = set()

    for row in ws.iter_rows(min_row=1):
        cell = row[0].value  # FIRST COLUMN
        if cell:
            hach_values.add(str(cell).strip().upper())

    wb.close()
    return hach_values
def get_first_payment_date(invoice_id):
    headers = {
        "Authorization": f"Zoho-oauthtoken {ACCESS_TOKEN or refresh_access_token()}",
        "X-com-zoho-inventory-organizationid": ORG_ID
    }

    try:
        payments_url = f"https://www.zohoapis.com/inventory/v1/invoices/{invoice_id}/payments"

        payments_resp = HTTP.get(payments_url, headers=headers)
        payments_resp.raise_for_status()

        payments = payments_resp.json().get("payments", [])

        if not payments:
            print(f"ℹ️ No payments found for invoice {invoice_id}")
            return None

        valid_dates = []

        for p in payments:
            raw_date = p.get("date")
            if raw_date:
                try:
                    parsed_date = datetime.strptime(raw_date[:10], "%Y-%m-%d")
                    valid_dates.append(parsed_date)
                except ValueError:
                    print(f"⚠️ Invalid payment date format: {raw_date}")

        if not valid_dates:
            print(f"ℹ️ No valid payment dates for invoice {invoice_id}")
            return None

        earliest_date = min(valid_dates)
        print(f"✅ Earliest payment date: {earliest_date.date()}")

        return earliest_date.strftime("%Y-%m-%d")

    except requests.exceptions.HTTPError as e:
        print(f"❌ HTTP error for invoice {invoice_id}: {e}")
        return None
    except Exception as e:
        print(f"❌ Failed to get payment date for {invoice_id}: {e}")
        return None
# =========== MAIN LOGIC ==========
def get_purchase_order_df(order_id: str) -> pd.DataFrame:
    # Get purchase order
    url = f"https://www.zohoapis.com/inventory/v1/purchaseorders/{order_id}"
    headers = {
        "Authorization": f"Zoho-oauthtoken {ACCESS_TOKEN or refresh_access_token()}",
        "X-com-zoho-inventory-organizationid": ORG_ID
    }
    
    response = HTTP.get(url, headers=headers)
    response.raise_for_status()
    po = response.json().get("purchaseorder", {})
    
    supplier = po.get("vendor_name")
    po_number = po.get("purchaseorder_number")
    date = po.get("date")
    reference = po.get("reference_number", "")
    
    if reference:
        reference = reference.strip("()").strip().rstrip(",")
    
    # Find SO numbers in reference (but proceed even if none found)
    so_numbers = re.findall(r"(?i)SO-\d+", reference) if reference else []
    so_info_by_sku = {}
    
    # Get sales orders if found
    if so_numbers:
        for so_num in so_numbers:
            so_num = so_num.upper()
            print(f"\nDebug: Fetching SO {so_num}")
            try:
                # First get the sales order to get its ID
                search_response = HTTP.get(
                    "https://www.zohoapis.com/inventory/v1/salesorders",
                    headers=headers,
                    params={"salesorder_number": so_num}
                )
                search_data = search_response.json()
                salesorders = search_data.get("salesorders", [])
                
                for so in salesorders:
                    if so.get("salesorder_number", "").upper() == so_num:
                        salesorder_id = so.get("salesorder_id")
                        # Now get the full sales order with line items
                        so_detail_url = f"https://www.zohoapis.com/inventory/v1/salesorders/{salesorder_id}"
                        so_response = HTTP.get(so_detail_url, headers=headers)
                        so_response.raise_for_status()
                        so_detail = so_response.json().get("salesorder", {})
                        delivery_condition = (so_detail.get("custom_field_hash", {}).get("cf_payment_conditions", ""))
                        invoices = so_detail.get("invoices", [])
                        delivery_cf = (
                            so_detail
                            .get("custom_field_hash", {})
                            .get("cf_delivery_after_payment", "")
                        )
                        delivery_date_range = None  # e.g. "10/02/2026 - 20/02/2026"
                        if isinstance(delivery_condition, str) and "after delivery" in delivery_condition.lower():
                            today = datetime.today().date()
                            match = re.search(r"(\d+)(?:\s*-\s*(\d+))?\s*(weeks?|კვირ\w*)", delivery_cf.lower())
                            if match:
                                start_w = int(match.group(1))
                                end_w = int(match.group(2)) if match.group(2) else start_w

                                start_date = today + timedelta(weeks=start_w)
                                end_date = today + timedelta(weeks=end_w)

                                start_str = start_date.strftime("%d/%m/%Y")
                                end_str = end_date.strftime("%d/%m/%Y")

                                delivery_date_range = (
                                    start_str
                                    if start_str == end_str
                                    else f"{start_str} - {end_str}"
                                )
                        elif isinstance(delivery_cf, str) and "ხელშეკრულებიდან" in delivery_cf:
                            range_match = re.search(r"(\d+)\s*[-–]\s*(\d+)\s*ხელშეკრულებიდან", delivery_cf)
                            single_match = re.search(r"(\d+)\s*ხელშეკრულებიდან", delivery_cf)
                            today = datetime.today().date()

                            if range_match:
                                min_days = int(range_match.group(1))
                                max_days = int(range_match.group(2))
                            elif single_match:
                                min_days = max_days = int(single_match.group(1))
                            else:
                                min_days = max_days = None
                            if min_days is not None:
                                start_date = today + timedelta(days=min_days)
                                end_date = today + timedelta(days=max_days)

                                if start_date == end_date:
                                    delivery_date_range = start_date.strftime('%d/%m/%Y')
                                else:
                                    delivery_date_range = f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
                        else:
                            try:
                                # Find the first paid/partially_paid invoice for this SO
                                target_invoice = None
                                for inv in invoices:
                                    status = inv.get("status", "").lower()
                                    if status in ["paid", "partially_paid"]:
                                        target_invoice = inv
                                        break
                                
                                # Process only the target invoice if found
                                if target_invoice:
                                    invoice_id = target_invoice.get('invoice_id')
                                    raw_payment_date = get_first_payment_date(invoice_id)
                                    if raw_payment_date:
                                        try:
                                            first_payment_date = datetime.strptime(raw_payment_date, "%Y-%m-%d").date()
                                        except ValueError:
                                            print(f"Debug: Unexpected payment date format: {raw_payment_date}")
                                            first_payment_date = None
                                    else:
                                        first_payment_date = None
                                    if first_payment_date:
                                        match = re.search(r"(\d+)(?:\s*-\s*(\d+))?\s*(weeks?|კვირ\w*)", delivery_cf.lower())

                                        if match:
                                            start_w = int(match.group(1))
                                            end_w = int(match.group(2)) if match.group(2) else start_w

                                            start_date = first_payment_date + timedelta(weeks=start_w)
                                            end_date = first_payment_date + timedelta(weeks=end_w)

                                            start_str = start_date.strftime("%d/%m/%Y")
                                            end_str = end_date.strftime("%d/%m/%Y")

                                            delivery_date_range = (
                                                start_str
                                                if start_str == end_str
                                                else f"{start_str} - {end_str}"
                                            )
                                        else:
                                            print(f"Debug: Delivery lead time format not recognized for SO {so_num}")
                                    else:
                                        print(f"Debug: Could not calculate delivery from payment for SO {so_num}")

                            except Exception as e:
                                print(f"Debug: Error fetching invoices for SO {so_num}: {e}")

                        line_items = so_detail.get("line_items", [])
                        # Process ALL line items - NO break here
                        for item in line_items:
                            sku = item.get("sku")
                            item_name = item.get("name")
                            if sku:
                                so_info_by_sku[sku] = {
                                    "SO": so_num,
                                    "SO_Customer": so_detail.get("customer_name"),
                                    "SO_Date": so_detail.get("date"),
                                    "SO_Status": so_detail.get("status"),
                                    "SO_Item_Name": item_name,
                                    "SO_Item_Quantity": item.get("quantity"),
                                    "SO_Country": so_detail.get("shipping_address", {}).get("country") 
                                    or so_detail.get("billing_address", {}).get("country") 
                                    or so_detail.get("country") 
                                    or "",
                                    "SO_Delivery_Date_Range": delivery_date_range
                                }
                        break  # Break only after processing this SO
                        
            except Exception as e:
                print(f"Debug: Error fetching SO {so_num}: {e}")
                continue
    
    # Debug: Print PO items
    for idx, item in enumerate(po.get("line_items", []), 1):
        sku = item.get("sku")
        matched = "Yes" if sku in so_info_by_sku else "No"
    
    # Create DataFrame - ALWAYS create for every PO
    items = []
    for item in po.get("line_items", []):
        sku = item.get("sku")
        so_data = so_info_by_sku.get(sku, {})
        is_match = "Yes" if sku in so_info_by_sku else "No"
        so_number = so_data.get("SO", "")
        po_customer = next((f.get("value_formatted")for f in item.get("item_custom_fields", []) if f.get("label") == "Customer" and f.get("value_formatted")),"")
        # Export logic for HACH
        export_value = ""
        if supplier == "HACH":
            country_lc = so_info_by_sku.get(sku, {}).get("SO_Country", "").lower()
            if "azerbaijan" in country_lc or "armenia" in country_lc:
                export_value = "კი"
            else:
                export_value = "არა"
        
        item_dict = {
            "Supplier Company": supplier,
            "PO": po_number,
            "შეკვეთის გაკეთების თარიღი": date,
            "Item": item.get("name"),
            "Code": sku,
            "Reference": reference,
            "შეკვეთილი რაოდენობა": item.get("quantity"),
            "Customer": po_customer if po_customer else so_data.get("SO_Customer", ""),
            "SO": so_number,
            "შეკვეთის ჩაბარების ვადა" : so_data.get("SO_Delivery_Date_Range", ""),
            "SO_Customer": so_data.get("SO_Customer", ""),
            "SO_Match": is_match,
            "Export?": export_value
        }
        items.append(item_dict)
    
    df = pd.DataFrame(items)
    
    # Print summary
    matches = df[df['SO_Match'] == 'Yes']
    print(f"SOs in reference: {', '.join(so_numbers) if so_numbers else 'None'}")
    print(f"Items matched: {len(matches)}/{len(df)}")
    # # Process HACH but still return DataFrame
    if supplier == "HACH":
        df = df.rename({"შეკვეთის ჩაბარების ვადა" : "მიწოდების ვადა"}, axis=1)
        process_hach(df)
    
    # ALWAYS return the DataFrame
    return df

def append_dataframe_to_table(df: pd.DataFrame, sheet_name: str):
    df = df[df['Supplier Company'] != 'HACH']
    perms_download = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{PERMS_ID}/content"
    headers = {"Authorization": f"Bearer {ACCESS_TOKEN_DRIVE or One_Drive_Auth()}"}

    max_attempts = 6
    for attempt in range(max_attempts):
        try:
            # --- Download permissions Excel file ---
            try:
                resp_perms = HTTP.get(perms_download, headers=headers, timeout=60)
                resp_perms.raise_for_status()
                perms_stream = io.BytesIO(resp_perms.content)
                perms_df = pd.read_excel(perms_stream, header=1)
                perms_stream.close()
                perms_stream = None
                if not {"მწარმოებლის კოდი", "მიღებული ნებართვა 1 / წერილის ნომერი"}.issubset(perms_df.columns):
                    print("⚠️ Warning: Permissions file missing required columns.")
                else:
                    perms_df = perms_df[["მწარმოებლის კოდი", "მიღებული ნებართვა 1 / წერილის ნომერი"]]

            except Exception as e_perm:
                print(f"⚠️ Could not download permissions Excel: {e_perm}")
                perms_df = pd.DataFrame(columns=["მწარმოებლის კოდი", "მიღებული ნებართვა 1 / წერილის ნომერი"])

            break  # success — exit retry loop

        except Exception as e:
            wait = min(5 * (attempt + 1), 30)
            print(f"⚠️ Error downloading main file (attempt {attempt+1}/{max_attempts}): {e}. Sleeping {wait}s")
            time.sleep(wait)
    else:
        print("❌ Gave up downloading files after multiple attempts")
        return

    items_df = pd.read_csv("zoho_items.csv")
    url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{TRANS_FILE}/content"
    resp = HTTP.get(url, headers=headers, timeout=60)
    resp.raise_for_status()
    trans = pd.read_excel(io.BytesIO(resp.content))
    trans_lookup = {}
    for _, row in trans.iterrows():
        if pd.notna(row['Item']) and pd.notna(row['თარგმანი']):
            normalized_item = str(row['Item']).translate(str.maketrans('', '', '.,\n\r\t')).lower().strip()
            trans_lookup[normalized_item] = row['თარგმანი']
    # Ensure table exists
    range_address = get_used_range(sheet_name)
    table_name = create_table_if_not_exists(range_address, sheet_name)
    # Handle Customer/Reference substitution
    if "Customer" in df.columns and "Reference" in df.columns:
        df = df.copy()
        for index, row in df.iterrows():
            customer_val = row['Customer']
            if (customer_val is None or 
                (isinstance(customer_val, str) and customer_val.strip() == "") or 
                (pd.isna(customer_val))):
                df.at[index, 'Customer'] = row['Reference']

        # ✅ Drop Reference column after substitution
        df = df.drop(columns=["Reference"])
    # Fetch table columns
    table_columns = get_table_columns(table_name)

    # Normalize DataFrame
    new_df = df.copy()
    for col in table_columns:
        if col not in new_df.columns:
            new_df[col] = ""
    new_df['#'] = range(1, len(new_df) + 1)

    # ✅ Restrict to table columns only
    out_df = new_df[table_columns]
    out_df["Code"] = out_df["Code"].astype(str).str.strip()

    items_df["sku"] = items_df["sku"].astype(str).str.strip()
    perms_df["მწარმოებლის კოდი"] = perms_df["მწარმოებლის კოდი"].astype(str).str.strip()
    # HS Code lookup: sku -> HS_Code
    hs_lookup = (
        items_df
        .drop_duplicates(subset="sku")
        .set_index("sku")["HS_Code"]
    )
    # Permission lookup: code -> letter
    perm_lookup = (
        perms_df
        .drop_duplicates(subset="მწარმოებლის კოდი")
        .set_index("მწარმოებლის კოდი")["მიღებული ნებართვა 1 / წერილის ნომერი"]
    )
    # Fill HS Code
    out_df["HS Code"] = out_df["Code"].map(hs_lookup)

    # Fill permissions
    out_df["წერილი"] = (
        out_df["Code"]
        .map(perm_lookup)
        .fillna("არ სჭირდება")
    )
    def get_translation(item):
        if pd.isna(item):
            return ""
        # Normalize the item text by removing punctuation
        normalized = str(item).translate(str.maketrans('', '', '.,\n\r\t')).lower().strip()
        return trans_lookup.get(normalized, "")
    
    out_df["თარგმანი"] = out_df["Item"].apply(get_translation)
    out_df['ადგილმდებარეობა'] = "წარმოების პროცესში"
    # --------------------------------------------------
    # 3️⃣ Final export
    # --------------------------------------------------
    rows = out_df.fillna("").astype(str).values.tolist()

    tbl_range_url = (
    f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}"
    f"/items/{FILE_ID}/workbook/tables/{table_name}/range"
    )
    headers = {"Authorization": f"Bearer {ACCESS_TOKEN_DRIVE}", "Content-Type": "application/json"}
    tbl_range = HTTP.get(tbl_range_url, headers=headers, timeout=30).json()["address"]

    tbl_range = tbl_range.split("!")[-1]  # A1:X57
    (start, end) = tbl_range.split(":")
    first_col, first_row = re.match(r"([A-Z]+)(\d+)", start).groups()
    last_col, last_row = re.match(r"([A-Z]+)(\d+)", end).groups()
    first_row, last_row = int(first_row), int(last_row)

    # ------------------ append rows ------------------
    url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{FILE_ID}/workbook/tables/{table_name}/rows/add"
    
    payload = {"values": rows}
    resp = HTTP.post(url, headers=headers, json=payload)

    if resp.status_code not in (200, 201):
        raise Exception(f"❌ Append failed: {resp.status_code} {resp.text[:200]}")

    print(f"✅ Appended {len(rows)} rows")

    # ------------------ color logic ------------------

    SUPPLIER_BASE_COLORS = {
        "KROHNE": (68,114,196), "Carl Roth": (255,0,0), "PENTAIR": (112,173,71),
        "In-Situ": (255,192,0), "VWR": (244,176,132), "Veolia Turkey": (192,0,0),
        "SAMSON": (172,185,202), "HYDROO": (255,192,0), "OTT HydroMet": (255,230,153),
        "Akkim": (155,194,230), "ATB WATER": (165,165,165), "ITM": (198,89,17),
        "AMAZON": (255,255,0), "STAR VALVE": (217,225,242),
        "VORTEX Water Engineering": (0,176,240), "KORHUS FILTER SYSTEMS": (172,185,202),
        "ToxSoft": (214,220,228), "NERO": (255,230,153), "AO Smith": (198,224,180)
    }
    SUPPLIER_BASE_COLORS_CI = {k.upper(): v for k, v in SUPPLIER_BASE_COLORS.items()}

    supplier_so_map = defaultdict(dict)
    supplier_so_counter = defaultdict(int)
    row_colors = []

    # Build colors for all rows in out_df (the ones just appended)
    for idx, r in out_df.iterrows():
        # --- supplier: column or fallback to second cell ---
        supplier = r.get("Supplier Company")
        if not supplier or str(supplier).strip() == "":
            supplier = r.iloc[1] if len(r) > 1 else ""
        
        supplier_key = str(supplier).strip().upper()  # normalize for case-insensitive lookup
        so = r.get("SO", "")
        
        # --- base color ---
        base = SUPPLIER_BASE_COLORS_CI.get(supplier_key, (220,220,220))
        
        # --- darker shade per SO ---
        if so not in supplier_so_map[supplier_key]:
            supplier_so_counter[supplier_key] += 1
            supplier_so_map[supplier_key][so] = supplier_so_counter[supplier_key]
        
        so_index = supplier_so_map[supplier_key][so]
        adjustment = (so_index - 1) * 35
        
        row_colors.append(
            tuple(max(0, min(255, int(c - adjustment))) for c in base)
        )

    # ------------------ Apply Colors to Excel ------------------
    # Determine first row of newly appended rows
    start_row = last_row + 1  # last_row is from your table range before append
    for i, (r,g,b) in enumerate(row_colors):
        row_idx = start_row + i
        rng = f"{first_col}{row_idx}:{last_col}{row_idx}"
        fill_url = (
            f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}"
            f"/items/{FILE_ID}/workbook/worksheets/{sheet_name}"
            f"/range(address='{rng}')/format/fill"
        )
        graph_safe_request( "PATCH", fill_url, headers, json={"color": f"#{r:02X}{g:02X}{b:02X}"}).raise_for_status()
    # ------------------ Apply Borders to All Appended Cells ------------------
    end_row = start_row + len(rows) - 1
    full_range = f"{first_col}{start_row}:{last_col}{end_row}"
    borders_url = (
        f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}"
        f"/items/{FILE_ID}/workbook/worksheets/{sheet_name}"
        f"/range(address='{full_range}')/format/borders"
    )
    border_payload = {
        "style": "Continuous",
        "weight": "Thin",
        "color": "#000000"
    }

    # Apply border to all edge types
    for border_type in ["EdgeTop","EdgeBottom","EdgeLeft","EdgeRight","InsideHorizontal","InsideVertical"]:
        graph_safe_request("PATCH", f"{borders_url}/{border_type}", headers, json=border_payload).raise_for_status()

def process_hach(df: pd.DataFrame) -> None:
    with EXCEL_LOCK:
        try:
            if df.empty:
                raise ValueError("Empty dataframe provided to process_hach")

            po_full = df["PO"].iloc[0]
            po_number = po_full.replace("PO-00", "")
            sheet_name = po_number

            print(f"\n📌 Creating HACH sheet '{sheet_name}'...")

            # ─────────────────────────────────────────────
            # 1️⃣ Base headers
            # ─────────────────────────────────────────────
            base_headers = {
                "Authorization": f"Bearer {ACCESS_TOKEN_DRIVE}",
                "Content-Type": "application/json"
            }

            # ─────────────────────────────────────────────
            # 2️⃣ Create workbook session (IMPORTANT)
            # ─────────────────────────────────────────────
            session = graph_safe_request(
                "POST",
                f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{HACH_FILE}/workbook/createSession",
                base_headers,
                {"persistChanges": True}
            ).json()

            session_headers = {
                **base_headers,
                "workbook-session-id": session["id"]
            }

            # ─────────────────────────────────────────────
            # 3️⃣ Add worksheet (INSIDE session)
            # ─────────────────────────────────────────────
            graph_safe_request(
                "POST",
                f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{HACH_FILE}/workbook/worksheets/add",
                session_headers,
                {"name": sheet_name}
            )

            
            # 2. Info table (must be exactly 4x2)
            info_data = [
                ["PO", po_number],
                ["SO", df["Reference"].iloc[0] if "Reference" in df else ""],
                ["POს გაკეთების თარიღი",pd.to_datetime(df["შეკვეთის გაკეთების თარიღი"].iloc[0]).strftime("%d/%m/%Y")],
                ["დღვანდელი თარიღი", pd.Timestamp.now().strftime("%d/%m/%Y")]
            ]

            graph_safe_request("PATCH",
                f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{HACH_FILE}"
                f"/workbook/worksheets/{sheet_name}/range(address='C3:D6')",
                session_headers,
                {"values": info_data}
            ).raise_for_status()

            # 3. Header row
            start_row = 8
            table_headers = [
                "Item", "წერილი", "Code", "HS Code", "Details", "თარგმანი", "QTY",
                "მიწოდების ვადა", "Confirmation 1 (shipment week)", "Packing List",
                "რა რიცხვში გამოგზავნეს Packing List-ი", "რამდენი გამოიგზავნა",
                "ჩამოსვლის სავარაუდო თარიღი", "რეალური ჩამოსვლის თარიღი",
                "Qty Delivered", "Customer", "Export?", "მდებარეობა", "შენიშვნა"
            ]

            write_range = f"B{start_row}:T{start_row}"

            graph_safe_request("PATCH",
                f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{HACH_FILE}"
                f"/workbook/worksheets/{sheet_name}/range(address='{write_range}')",
                session_headers,
                {"values": [table_headers]}
            ).raise_for_status()

            # 4. Create MS Graph Table
            table_resp = graph_safe_request("POST",
                f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{HACH_FILE}/workbook/tables/add",
                session_headers,
                {"address": f"{sheet_name}!{write_range}", "hasHeaders": True}
            )
            table_resp.raise_for_status()

            table_id = table_resp.json()["id"]

            # 5. Add rows in batches
            normalized_df = normalize_hach(df)
            rows = normalized_df.values.tolist()

            batch_size = 50
            for i in range(0, len(rows), batch_size):
                batch = rows[i:i + batch_size]

                r = graph_safe_request("POST",
                    f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{HACH_FILE}"
                    f"/workbook/tables/{table_id}/rows/add",
                    session_headers,
                    {"values": batch}
                )
                r.raise_for_status()

                print(f"   ➕ Added batch {i // batch_size + 1}")

            print(f"✅ HACH workflow completed. Added {len(rows)} rows.")
            POOL.submit(format_hach_sheet_full, sheet_name,start_row, len(normalized_df), table_id)

        except Exception as e:
            print(f"❌ HACH processing failed: {e}")
            traceback.print_exc()
            raise

def process_shipment(order_number: str, items: list) -> None:
        try:
            # --- Load sheet values ---
            data = get_sheet_values("მიმდინარე ")

            # Ensure proper row formatting
            data = [list(row) for row in data]

            # Build DataFrame safely
            df_source = pd.DataFrame(data[1:], columns=data[0])
            df_source["Code"] = df_source["Code"].astype(str).str.strip()
            df_source["შეკვეთილი რაოდენობა"] = df_source["შეკვეთილი რაოდენობა"]

            # --- Filter matching rows ---
            order_number = str(order_number).strip()
            matching = df_source[df_source["SO"].astype(str).str.strip() == order_number].copy()


            if matching.empty:
                print(f"⚠️ No rows found for SO = {order_number}")
                return
            delivered_by_sku = defaultdict(float)
            for item in items:
                sku = item["sku"].strip().upper()
                delivered_by_sku[sku] += float(item.get("quantity", 0))
            rows_to_move = []

            for idx, row in matching.iterrows():
                sku = row["Code"].strip().upper()
                qty_ordered = float(row["შეკვეთილი რაოდენობა"])
                qty_delivered_so_far = float(row.get("მიწოდებული რაოდენობა", 0))

                # Delivered in this package
                newly_delivered = delivered_by_sku.get(sku, 0)

                if newly_delivered == 0:
                    continue

                total_delivered = qty_delivered_so_far + newly_delivered

                # Update delivered quantity in source DF
                df_source.at[idx, "მიწოდებული რაოდენობა"] = total_delivered

                if total_delivered >= qty_ordered:
                    rows_to_move.append(idx)
                else:
                    print(
                        f"⏳ Partial delivery: SO={order_number}, "
                        f"SKU={sku}, Ordered={qty_ordered}, "
                        f"Delivered={total_delivered}"
                    )

            # --- Prepare DataFrame to move ---
            df_move = matching.loc[rows_to_move].copy()
            df_move["ადგილმდებარეობა"] = "ჩაბარდა"

            # --- Append to destination sheet ---
            append_dataframe_to_table(df_move, "ჩამოსული")

            # --- Delete moved rows from source ---
            start_row = get_table_start_row_from_used_range("მიმდინარე ")
            worksheet_rows = [start_row + 1 + idx for idx in rows_to_move]
            delete_table_rows("მიმდინარე ", worksheet_rows)

            print(f"✅ Completed processing for SO {order_number}, moved {len(rows_to_move)} rows")

        except Exception as e:
            print(f"❌ Fatal error: {e}")
            traceback.print_exc()

def recieved_hach(po_number: str,date:str, items: list[dict]) -> None:
    po_sheet = re.sub(r"\D", "", po_number).lstrip("00")
    print(f"📄 HACH sheet name: {po_sheet}")
    with EXCEL_LOCK:
        headers = {"Authorization": f"Bearer {ACCESS_TOKEN_DRIVE or One_Drive_Auth()}"}
        url_download = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{HACH_FILE}/content"
        resp = HTTP.get(url_download, headers=headers, timeout=60)
        resp.raise_for_status()
        file_stream = io.BytesIO(resp.content)
        wb = load_workbook(file_stream)
        
        if po_sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{po_sheet}' not found in HACH file")

        ws = wb[po_sheet]

        # --- Get first table ---
        if not ws.tables:
            raise ValueError(f"No tables found in sheet '{po_sheet}'")
        tables = list(ws.tables.values())
        table = tables[0]
        start_cell, end_cell = table.ref.split(":")
        start_row = ws[start_cell].row
        start_col = ws[start_cell].column
        end_row = ws[end_cell].row
        end_col = ws[end_cell].column

        print(f"📊 Using table {table.name} ({table.ref})")

        # --- Read table into pandas ---
        data = [
        list(r) for r in ws.iter_rows(
            min_row=start_row,
            max_row=end_row,
            min_col=start_col,
            max_col=end_col,
            values_only=True
        )
        ]

        df = pd.DataFrame(data[1:], columns=data[0])
        df["რეალური ჩამოსვლის თარიღი"] = (pd.to_datetime(date) - pd.Timedelta(days=2)).date()
        df['მდებარეობა'] = "ოფისი"

        df["Details"] = df["Details"].astype(str).str.strip()

        pr_items = []

        for item in items:
            name = str(item.get("name")).strip()
            if not name:
                continue

            qty = item.get("quantity") or 0
            try:
                qty = float(qty)
            except (TypeError, ValueError):
                qty = 0

            pr_items.append({
                "name": name,
                "quantity": qty,
                "used": False
            })

        print("📦 Purchase Receive items (ordered):")
        for i in pr_items:
            print(f"   {i['name']} → {i['quantity']}")

        # --- Fill Qty Delivered based on Details / item_name ---
        updated = 0
        for idx, row in df.iterrows():
            details_norm = str(row["Details"]).strip()

            # find first unused PR item with same name
            for pr in pr_items:
                if not pr["used"] and pr["name"] == details_norm:
                    df.at[idx, "Qty Delivered"] = pr["quantity"]
                    pr["used"] = True
                    updated += 1
                    print(f"   ✔ {row['Details']} → {pr['quantity']}")
                    break

        if updated == 0:
            print("⚠️ No items matched Excel Details column")
            return
        mask_coo = df["Code"] == "CoO"
        df.loc[mask_coo, "მდებარეობა"] = df.loc[mask_coo, "მდებარეობა"].bfill()
        df.loc[mask_coo, "რეალური ჩამოსვლის თარიღი"] = df.loc[mask_coo, "რეალური ჩამოსვლის თარიღი"].bfill()
        # --- Write back to Excel ---
        for r_idx, row in enumerate(df.values.tolist(), start=start_row + 1):
            for c_idx, value in enumerate(row, start=start_col):
                ws.cell(row=r_idx, column=c_idx).value = value
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        upload_url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{HACH_FILE}/content"

        for attempt in range(8):
            resp = HTTP.put(upload_url, headers=headers, data=output.getvalue())
            if resp.status_code in (409, 423):
                time.sleep(min(30, 2 ** attempt))
                continue
            resp.raise_for_status()
            print(f"✅ Packing List updated successfully ({updated} rows)")
            return

def recieved_nonhach(po_number: str, date:str, line_items: list[dict]) -> None:
    with EXCEL_LOCK:
        file_stream = None
        wb = None

        try:
            # --- Step 0: Build PR items (ORDER PRESERVING) ---
            po_str = str(po_number).strip()
            pr_items = []

            for item in line_items:
                name = item.get("name")
                qty = item.get("quantity")

                if not name:
                    continue

                try:
                    qty = float(qty)
                except (TypeError, ValueError):
                    qty = 0

                pr_items.append({
                    "po": po_str,
                    "name": str(name).strip().lower(),
                    "quantity": qty,
                    "used": False
                })

            if not pr_items:
                print("⚠️ No valid PR items to process")
                return

            print("📦 Incoming Purchase Receive items:")
            for p in pr_items:
                print(f"   {p['po']} | {p['name']} → {p['quantity']}")

            # --- Step 1: Download Excel ---
            url_download = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{FILE_ID}/content"
            headers = {"Authorization": f"Bearer {ACCESS_TOKEN_DRIVE or One_Drive_Auth()}"}

            for attempt in range(6):
                try:
                    resp = HTTP.get(url_download, headers=headers, timeout=60)
                    resp.raise_for_status()
                    file_stream = io.BytesIO(resp.content)
                    wb = load_workbook(file_stream)
                    break
                except Exception as e:
                    wait = min(5 * (attempt + 1), 30)
                    print(f"⚠️ Download failed ({attempt+1}/6): {e}, retrying in {wait}s")
                    time.sleep(wait)
            else:
                print("❌ Failed to download Excel")
                return

            # --- Step 2: Choose target sheet based on PO ---
            target_sheet = None
            target_df = None

            for sheet_name in ("მიმდინარე ", "ჩამოსული"):
                ws = wb[sheet_name]
                df = pd.DataFrame(ws.values)
                df.columns = df.iloc[0]
                df = df[1:]

                df["PO"] = df["PO"].astype(str).str.strip()

                if (df["PO"] == po_str).any():
                    target_sheet = sheet_name
                    target_df = df
                    print(f"📄 Using sheet '{sheet_name}'")
                    break

            if target_sheet is None:
                print(f"⚠️ PO {po_str} not found in any sheet")
                return

            ws = wb[target_sheet]

            # --- Step 3: Validate & normalize ---
            required_cols = {"PO", "Item", "რეალურად გამოგზავნილი რაოდენობა"}
            if not required_cols.issubset(target_df.columns):
                raise ValueError(f"Missing required columns in '{target_sheet}'")

            target_df["Item"] = (
                target_df["Item"]
                .astype(str)
                .str.strip()
                .str.lower()
            )
            po_mask = target_df["PO"] == po_str
            target_df.loc[po_mask, "ჩამოსვლის თარიღი"] = (pd.to_datetime(date) - pd.Timedelta(days=2)).date()
            # --- Step 4: Order-preserving fill ---
            updated = 0

            for idx, row in target_df.iterrows():
                for pr in pr_items:
                    if (
                        not pr["used"]
                        and row["PO"] == pr["po"]
                        and row["Item"] == pr["name"]
                    ):
                        current_qty = row.get("რეალურად გამოგზავნილი რაოდენობა")
                        if current_qty in (None, ""):
                            target_df.at[idx, "რეალურად გამოგზავნილი რაოდენობა"] = pr["quantity"]
                            pr["used"] = True
                            updated += 1
                            print(f"   ✔ {row['Item']} → {pr['quantity']}")
                        else:
                            print(f"   ℹ️ Skipped {row['Item']}, cell already has value {current_qty}")
                        break

            if updated == 0:
                print("⚠️ No rows updated")
                return

            print(f"✅ Updated {updated} rows in '{target_sheet}'")

            # --- Step 5: Write back to Excel ---
            for col_idx, col_name in enumerate(target_df.columns, start=1):
                ws.cell(row=1, column=col_idx).value = col_name

            for row_idx, row in enumerate(target_df.values.tolist(), start=2):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value in (None, ""):
                        cell.value = value

            # --- Step 6: Save & upload ---
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            url_upload = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{FILE_ID}/content"

            for attempt in range(10):
                resp = HTTP.put(url_upload, headers=headers, data=output.getvalue())
                if resp.status_code in (423, 409):
                    wait = min(30, 2 ** attempt)
                    print(f"⚠️ File locked, retrying in {wait}s")
                    time.sleep(wait)
                    continue

                resp.raise_for_status()
                print("✅ Excel upload successful")
                return

            raise RuntimeError("Upload failed after retries")

        except Exception as e:
            print(f"❌ Fatal error: {e}")

        finally:
            if wb:
                wb.close()
            if file_stream:
                file_stream.close()
            gc.collect()

def process_message(mailbox, message_id, message_date, internet_id):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    try:
        cursor.execute(
            "INSERT INTO processed_messages (internet_id) VALUES (?)",
            (internet_id,)
        )
        conn.commit()
    except sqlite3.IntegrityError:
        print("⚠️ Duplicate email skipped")
        conn.close()
        return
    print(f"Mailbox: {mailbox}")
    print(f"message_id: {message_id}")
    print(f"message_date: {message_date}")
    att_url = f"https://graph.microsoft.com/v1.0/users/{mailbox}/messages/{message_id}/attachments"
    att_resp = HTTP.get(att_url, headers=get_headers(), timeout=20)
    if att_resp.status_code != 200:
        print(f"❌ Error fetching attachments: {att_resp.status_code} - {att_resp.text}")
        return
    attachments = att_resp.json().get("value", [])
    po_text_map = {}

    pdf_attachments = [
        att for att in attachments
        if att.get("name", "").lower().endswith(".pdf")
        or att.get("contentType") == "application/pdf"
    ]

    if not pdf_attachments:
        print("ℹ️ No PDF attachments found")
        return

    for att in pdf_attachments:

        if "contentBytes" not in att:
            print(f"❌ Attachment {att.get('name')} has no contentBytes - skipping")
            continue

        print(f"📎 Processing PDF: {att.get('name')}")

        content = base64.b64decode(att["contentBytes"])
        all_text = ""

        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                all_text += (page.extract_text() or "") + "\n"

        po_match = re.search(r"PO-\d+", all_text)
        po_number = po_match.group(0) if po_match else None

        if not po_number:
            print(f"⚠️ No PO found in {att.get('name')} → skipping")
            continue

        po_text_map[po_number] = all_text
        print(f"🎯 Found PO {po_number}")
    if isinstance(message_date, str):
        dt = datetime.fromisoformat(message_date.replace("Z", "+00:00"))
    elif isinstance(message_date, datetime):
        dt = message_date
    else:
        print(f"⚠️ Unexpected message_date type: {type(message_date)}")
        return

    confirmation_date = dt.date()
    with EXCEL_LOCK:
        file_stream = None
        wb = None
        orders_df = pd.DataFrame()

        # --- Step 1: Download current orders Excel file ---
        url_download = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{FILE_ID}/content"
        headers = {"Authorization": f"Bearer {ACCESS_TOKEN_DRIVE or One_Drive_Auth()}"}

        max_attempts = 6
        for attempt in range(max_attempts):
            try:
                # --- Download orders file ---
                resp = HTTP.get(url_download, headers=headers, timeout=60)
                resp.raise_for_status()
                file_stream = io.BytesIO(resp.content)
                wb = load_workbook(file_stream)

                if "მიმდინარე " in wb.sheetnames:
                    ws = wb["მიმდინარე "]
                    orders_df = pd.DataFrame(ws.values)
                    orders_df.columns = orders_df.iloc[0]  # first row as header
                    orders_df = orders_df[1:]              # drop header row
                else:
                    print("⚠️ Worksheet 'მიმდინარე ' not found in orders file.")
                    orders_df = pd.DataFrame()
                break  # success — exit retry loop

            except Exception as e:
                wait = min(5 * (attempt + 1), 30)
                print(f"⚠️ Error downloading main file (attempt {attempt+1}/{max_attempts}): {e}. Sleeping {wait}s")
                time.sleep(wait)

        else:
            print("❌ Gave up downloading files after multiple attempts")
            return

        if not po_text_map:
            print("ℹ️ No PO found in any PDF → skipping Excel update")
            return

        updated_rows = 0

        for po_number, all_text in po_text_map.items():

            print(f"\n🔄 Updating Excel for PO {po_number}")

            matching_idx = orders_df.index[orders_df["PO"] == po_number]

            if len(matching_idx) == 0:
                print(f"⚠️ No Excel rows found for PO {po_number}")
                continue

            for idx in matching_idx:

                code = str(orders_df.at[idx, "Code"]).strip()
                print(f"🔍 Checking code: {code}")

                if code and code in all_text:
                    print(f"✅ Code {code} found in PDF")

                    current_val = orders_df.at[idx, "Confirmation-ის მოსვლის თარიღი"]

                    if pd.isna(current_val) or current_val == "":
                        orders_df.at[idx, "Confirmation-ის მოსვლის თარიღი"] = confirmation_date
                        updated_rows += 1
                        print("   Filled confirmation date")

        if updated_rows == 0:
            print("⚠️ No matching item codes found in any confirmation PDFs.")
            return

        # 🟢 after loop, update sheet once:
        ws = wb["მიმდინარე "]

        # Write headers if needed
        for col_idx, col_name in enumerate(orders_df.columns.tolist(), start=1):
            ws.cell(row=1, column=col_idx).value = col_name

        # Write data values
        for row_idx, row in enumerate(orders_df.values.tolist(), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                if orders_df.columns[col_idx - 1] == "Confirmation-ის მოსვლის თარიღი" and value:
                    cell.number_format = "DD/MM/YYYY"

        # Save workbook to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Upload back
        url_upload = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{FILE_ID}/content"
        max_attempts = 10
        for attempt in range(max_attempts):
            resp = HTTP.put(url_upload, headers=headers, data=output.getvalue())
            if resp.status_code in (423, 409):  # Locked
                wait_time = min(30, 2**attempt) + random.uniform(0, 2)
                print(f"⚠️ File locked (attempt {attempt+1}/{max_attempts}), retrying in {wait_time:.1f}s...")
                time.sleep(wait_time)
                continue

            resp.raise_for_status()
            range_address = get_used_range("მიმდინარე ")
            table_name = create_table_if_not_exists(range_address, "მიმდინარე ")
            print(f"✅ Upload successful. Created table named {table_name}")
            file_stream.close()
            file_stream = wb = None
            del orders_df
            gc.collect()
            return
    conn.close()

def process_hach_message(mailbox, message_id, message_date, internet_id):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    try:
        cursor.execute(
            "INSERT INTO processed_messages (internet_id) VALUES (?)",
            (internet_id,)
        )
        conn.commit()
    except sqlite3.IntegrityError:
        print("⚠️ Duplicate email skipped")
        conn.close()
        return
    print(f"📦 HACH processing | mailbox={mailbox}, message_id={message_id}")
    if isinstance(message_date, str):
        dt = datetime.fromisoformat(message_date.replace("Z", "+00:00"))
    elif isinstance(message_date, datetime):
        dt = message_date
    else:
        print(f"⚠️ Unexpected message_date type: {type(message_date)}")
        return

    confirmation_date = dt.date()  # <-- DATE ONLY

    with EXCEL_LOCK:
        headers = {"Authorization": f"Bearer {ACCESS_TOKEN_DRIVE or One_Drive_Auth()}"}

        # --------------------------------------------------
        # 1. Fetch message → subject (PO number)
        # --------------------------------------------------
        msg_url = f"https://graph.microsoft.com/v1.0/users/{mailbox}/messages/{message_id}"
        msg_resp = HTTP.get(msg_url, headers=headers, timeout=20)
        msg_resp.raise_for_status()
        message = msg_resp.json()

        subject = message.get("subject", "").strip()

        po_match = re.search(r"\bPO[-:#–]?\s*(\d+)\b", subject, re.IGNORECASE)
        if not po_match:
            print(f"❌ No PO number found in subject: {subject!r}")
            return

        sheet_name = str(int(po_match.group(1)))
        print(f"📄 Target sheet extracted from subject: {sheet_name}")

        # --------------------------------------------------
        # 2. Download Excel files
        # --------------------------------------------------
        def download_excel(file_id):
            url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{file_id}/content"
            resp = HTTP.get(url, headers=headers, timeout=60)
            resp.raise_for_status()
            return io.BytesIO(resp.content)
        main_stream   = download_excel(HACH_FILE)

        wb = load_workbook(main_stream)

        ws = wb[sheet_name]

        # --------------------------------------------------
        # 3. Extract the ONLY table in the sheet
        # --------------------------------------------------
        tables = list(ws.tables.values())
        if len(tables) != 1:
            print(f"❌ Expected exactly 1 table, found {len(tables)}")
            return

        table = tables[0]
        start_cell, end_cell = table.ref.split(":")
        start_row = ws[start_cell].row
        start_col = ws[start_cell].column
        end_row   = ws[end_cell].row
        end_col   = ws[end_cell].column

        data = [
            list(r) for r in ws.iter_rows(
                min_row=start_row,
                max_row=end_row,
                min_col=start_col,
                max_col=end_col,
                values_only=True
            )
        ]

        df = pd.DataFrame(data[1:], columns=data[0])
        df["Code"] = df["Code"].astype(str).str.strip()

        # --------------------------------------------------
        # 5. Fetch and parse PDF
        # --------------------------------------------------
        att_url = f"https://graph.microsoft.com/v1.0/users/{mailbox}/messages/{message_id}/attachments"
        att_resp = HTTP.get(att_url, headers=headers, timeout=20)
        att_resp.raise_for_status()

        pdfs = [
            a for a in att_resp.json().get("value", [])
            if a.get("name", "").lower().endswith(".pdf")
        ]

        if len(pdfs) != 1:
            print(f"❌ Expected 1 PDF, found {len(pdfs)}")
            return

        content = base64.b64decode(pdfs[0]["contentBytes"])

        pdf_text = ""
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                pdf_text += (page.extract_text(layout=True) or "") + "\n"

        code_week_map = {}

        # First, extract all week information with their context
        week_pattern = re.compile(r'Dispatch\s+in\s+week\s*:\s*(\d{1,2}/\d{4})', re.IGNORECASE)
        
        # Find all occurrences of weeks and their positions
        week_matches = list(week_pattern.finditer(pdf_text))
        
        if week_matches:
            # For each code in the dataframe, search for it in the PDF
            for idx, row in df.iterrows():
                excel_code = str(row["Code"]).strip()
                if not excel_code:
                    continue
                
                # Search for this code in the PDF
                code_pattern = re.compile(r'\b' + re.escape(excel_code) + r'\b')
                code_match = code_pattern.search(pdf_text)
                
                if code_match:
                    # Find the nearest week after this code
                    code_pos = code_match.end()
                    nearest_week = None
                    
                    for week_match in week_matches:
                        week_pos = week_match.start()
                        # Find the first week that appears after the code
                        if week_pos > code_pos:
                            nearest_week = week_match.group(1)
                            break
                    
                    if nearest_week and excel_code not in code_week_map:
                        code_week_map[excel_code] = nearest_week

        if not code_week_map:
            print("⚠️ No code-week pairs found in PDF")

        # --------------------------------------------------
        # 7. Update rows by Code
        # --------------------------------------------------
        updated = 0

        for idx, row in df.iterrows():
            code = str(row["Code"]).strip().upper()

            if not code or code not in code_week_map:
                continue

            week_number = code_week_map[code]
            current_val = row.get("Confirmation 1 (shipment week)")
            if is_empty(current_val):
                df.at[idx, "Confirmation 1 (shipment week)"] = (
                    f"{confirmation_date.strftime('%d.%m.%Y')} (week {week_number})"
                )
                updated += 1
            else:
                print(f"ℹ️ Skipped overwrite for code {code} (already has value)")

            updated += 1

        if updated == 0:
            print("⚠️ No codes from PDF matched table")

        # --------------------------------------------------
        # 7. Write table back to sheet
        # --------------------------------------------------

        # For rows where Code == "CoO", backward fill from the next non-NaN
        mask_coo = df["Code"] == "CoO"
        df.loc[mask_coo, "Confirmation 1 (shipment week)"] = df.loc[mask_coo, "Confirmation 1 (shipment week)"].bfill()
        for r_idx, row in enumerate(df.values.tolist(), start=start_row + 1):
            for c_idx, value in enumerate(row, start=start_col):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                if df.columns[c_idx - start_col] == "Confirmation 1 (shipment week)" and value:
                    cell.number_format = "DD/MM/YYYY"

        # --------------------------------------------------
        # 8. Upload updated workbook
        # --------------------------------------------------
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        upload_url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{HACH_FILE}/content"

        for attempt in range(8):
            resp = HTTP.put(upload_url, headers=headers, data=output.getvalue())
            if resp.status_code in (409, 423):
                time.sleep(min(30, 2 ** attempt))
                continue
            resp.raise_for_status()
            print(f"✅ HACH update successful ({updated} rows)")
            return
    conn.close()

def process_khrone_message(mailbox, message_id, message_date, internet_id):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    try:
            cursor.execute(
                "INSERT INTO processed_messages (internet_id) VALUES (?)",
                (internet_id,)
            )
            conn.commit()
    except sqlite3.IntegrityError:
        print("⚠️ Duplicate email skipped")
        conn.close()
        return
    print(f"Mailbox: {mailbox}")
    print(f"message_id: {message_id}")
    print(f"message_date: {message_date}")

    if isinstance(message_date, str):
        dt = datetime.fromisoformat(message_date.replace("Z", "+00:00"))
    elif isinstance(message_date, datetime):
        dt = message_date
    else:
        print(f"⚠️ Unexpected message_date type: {type(message_date)}")
        return

    confirmation_date = dt.date()  # <-- DATE ONLY

    # --- Download attachments ---
    att_url = f"https://graph.microsoft.com/v1.0/users/{mailbox}/messages/{message_id}/attachments"
    att_resp = HTTP.get(att_url, headers=get_headers(), timeout=20)
    if att_resp.status_code != 200:
        print(f"❌ Error fetching attachments: {att_resp.status_code} - {att_resp.text}")
        return

    attachments = att_resp.json().get("value", [])
    pdf_attachments = [
        att for att in attachments
        if att.get("name", "").upper().startswith("SO_")
        and att.get("name", "").lower().endswith(".pdf")
    ]

    if not pdf_attachments:
        print("⚠️ No matching SO_ PDF attachments found")
        return
    item_results_by_po = {}
    for att in pdf_attachments:
        print(f"📎 Processing attachment: {att['name']}")

        content = base64.b64decode(att['contentBytes'])

        with pdfplumber.open(io.BytesIO(content)) as pdf:
            pdf_text_pages = []

            for page in pdf.pages:
                text = page.extract_text() or ""
                text = re.sub(r'\s+', ' ', text)
                pdf_text_pages.append(text)

        # --- Extract PO from THIS PDF ---
        po_number = None
        for page_text in pdf_text_pages:
            po_match = re.search(r"PO-\d+", page_text)
            if po_match:
                po_number = po_match.group(0)
                break

        if not po_number:
            print(f"❌ PO number not found in {att['name']}, skipping")
            continue

        print(f"🎯 Found PO: {po_number}")

        # --- Extract items for THIS PDF ---
        item_week_pattern = re.compile(r'([A-Z0-9]{10,})\s*Week\s*(\d{1,2}/\d{4})', re.IGNORECASE)
        quantity_pattern = re.compile(r'(\d+)\s*pcs', re.IGNORECASE)

        items_info = {}

        for page_text in pdf_text_pages:
            page_text = re.sub(r'\s+', ' ', page_text)

            for match in item_week_pattern.finditer(page_text):
                code = match.group(1)
                week = match.group(2)

                start_pos = max(match.start() - 100, 0)
                qty_search = page_text[start_pos:match.start()]

                qty_matches = list(quantity_pattern.finditer(qty_search))
                quantity = int(qty_matches[-1].group(1)) if qty_matches else None

                items_info[code] = {
                    "week": week,
                    "quantity": quantity
                }

        item_results_by_po[po_number] = items_info

    print(item_results_by_po)
       
    if not items_info:
        print("⚠️ No item codes found in PDFs")
        return

    # --- Update Excel only for rows with this PO ---
    with EXCEL_LOCK:
        file_stream = None
        wb = None
        orders_df = pd.DataFrame()

        url_download = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{FILE_ID}/content"
        headers = {"Authorization": f"Bearer {ACCESS_TOKEN_DRIVE or One_Drive_Auth()}"}

        max_attempts = 6
        for attempt in range(max_attempts):
            try:
                # --- Download orders file ---
                resp = HTTP.get(url_download, headers=headers, timeout=60)
                resp.raise_for_status()
                file_stream = io.BytesIO(resp.content)
                wb = load_workbook(file_stream)

                if "მიმდინარე " in wb.sheetnames:
                    ws = wb["მიმდინარე "]
                    orders_df = pd.DataFrame(ws.values)
                    orders_df.columns = orders_df.iloc[0]  # first row as header
                    orders_df = orders_df[1:].reset_index(drop=True)
                    orders_df["_excel_row"] = range(2, 2 + len(orders_df))  # Excel rows start at 2
                else:
                    print("⚠️ Worksheet 'მიმდინარე ' not found in orders file.")
                    orders_df = pd.DataFrame()
                break  # success — exit retry loop

            except Exception as e:
                wait = min(5 * (attempt + 1), 30)
                print(f"⚠️ Error downloading main file (attempt {attempt+1}/{max_attempts}): {e}. Sleeping {wait}s")
                time.sleep(wait)

        else:
            print("❌ Gave up downloading files after multiple attempts")
            return
        updated_rows = 0

        for po_number, items_info in item_results_by_po.items():

            po_rows = orders_df[orders_df["PO"] == po_number]

            if po_rows.empty:
                print(f"⚠️ No Excel rows found for PO {po_number}")
                continue

            for i, row in po_rows.iterrows():

                code = str(row["Code"]).strip()

                if code not in items_info:
                    print(f"⚠️ Code {code} not found in PDF for PO {po_number}")
                    continue

                info = items_info[code]

                # --- Update week ---
                current_week = orders_df.at[i, "Confirmation-ის მოსვლის თარიღი"]

                if is_empty(current_week):
                    orders_df.at[i, "Confirmation-ის მოსვლის თარიღი"] = \
                        f"{confirmation_date} (Week {info['week']})"

                # --- Update quantity ---
                if info["quantity"] is not None:
                    orders_df.at[i, "რეალურად გამოგზავნილი რაოდენობა"] = info["quantity"]

                updated_rows += 1
                print(f"✅ Updated PO {po_number} | Code {code}")

        if updated_rows == 0:
            print("⚠️ No rows were updated")

        # =====================================================
        # ⭐ Write Back To Excel
        # =====================================================

        for i, row in orders_df.iterrows():
            excel_row = row["_excel_row"]

            for col_idx, col_name in enumerate(orders_df.columns, start=1):
                if col_name == "_excel_row":
                    continue

                ws.cell(row=excel_row, column=col_idx).value = row[col_name]

                if col_name == "Confirmation-ის მოსვლის თარიღი" and row[col_name]:
                    ws.cell(row=excel_row, column=col_idx).number_format = "DD/MM/YYYY"

    # Save workbook to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Upload back
    url_upload = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{FILE_ID}/content"
    max_attempts = 10
    for attempt in range(max_attempts):
        resp = HTTP.put(url_upload, headers=headers, data=output.getvalue())
        if resp.status_code in (423, 409):  # Locked
            wait_time = min(30, 2**attempt) + random.uniform(0, 2)
            print(f"⚠️ File locked (attempt {attempt+1}/{max_attempts}), retrying in {wait_time:.1f}s...")
            time.sleep(wait_time)
            continue

        resp.raise_for_status()
        range_address = get_used_range("მიმდინარე ")
        table_name = create_table_if_not_exists(range_address, "მიმდინარე ")
        print(f"✅ Upload successful. Created table named {table_name}")
        file_stream.close()
        file_stream = wb = None
        del orders_df
        gc.collect()
        return
    conn.close()

def packing_list(mailbox, message_id, message_date, internet_id):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    try:
        cursor.execute(
            "INSERT INTO processed_messages (internet_id) VALUES (?)",
            (internet_id,)
        )
        conn.commit()
    except sqlite3.IntegrityError:
        print("⚠️ Duplicate email skipped")
        conn.close()
        return
    print(f"📦 Packing List processing | mailbox={mailbox}, message_id={message_id}")

    with EXCEL_LOCK:
        headers = {"Authorization": f"Bearer {ACCESS_TOKEN_DRIVE or One_Drive_Auth()}"}

        # --- Step 1: Fetch message metadata ---
        msg_url = f"https://graph.microsoft.com/v1.0/users/{mailbox}/messages/{message_id}"
        msg_resp = HTTP.get(msg_url, headers=headers, timeout=20)
        msg_resp.raise_for_status()
        message = msg_resp.json()

        subject = message.get("subject", "").strip()
        k_numbers = re.findall(r"K\d+", subject, re.IGNORECASE)
        if not k_numbers:
            print(f"❌ No K numbers found in subject: {subject!r}")
            return

        k_numbers = [k.upper() for k in k_numbers]
        print(f"📦 Found Packing Lists in subject: {k_numbers}")

        if isinstance(message_date, str):
            dt = datetime.fromisoformat(message_date.replace("Z", "+00:00"))
        else:
            dt = message_date

        confirmation_date_str = dt.strftime("%d/%m/%Y")
        arrival_date_str = (dt + timedelta(weeks=3)).strftime("%d/%m/%Y")

        # --- Step 2: Fetch attachments ---
        att_url = f"https://graph.microsoft.com/v1.0/users/{mailbox}/messages/{message_id}/attachments"
        att_resp = HTTP.get(att_url, headers=headers, timeout=20)
        att_resp.raise_for_status()
        attachments = att_resp.json().get("value", [])
        file_pattern = re.compile(r"^GG\w+$", re.IGNORECASE)

        pdf_attachments = [
            a for a in attachments
            if file_pattern.match(a['name'].split(".")[0])
            and a['name'].lower().endswith((".pdf", ".rtf"))
        ]

        file_bytes = base64.b64decode(pdf_attachments[0]['contentBytes'])
        # --- Step 3: Extract text ---
        pdf_text = ""

        if pdf_attachments[0]['name'].lower().endswith(".pdf"):
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                for page in pdf.pages:
                    chars = sorted(page.chars, key=lambda c: (c['top'], c['x0']))
                    current_line, last_top, last_x = [], None, None

                    for c in chars:
                        if last_top is None or abs(c['top'] - last_top) > 3:
                            if current_line:
                                pdf_text += "".join(current_line) + "\n"
                            current_line = [c['text']]
                            last_top, last_x = c['top'], c['x1']
                        else:
                            if c['x0'] - last_x > 2:
                                current_line.append(" ")
                            current_line.append(c['text'])
                            last_x = c['x1']

                    if current_line:
                        pdf_text += "".join(current_line) + "\n"
        else:
            pdf_text = file_bytes.decode(errors="ignore")

        # --- Step 4: Extract ALL PO numbers ---
        po_numbers = [
            str(int(m)) for m in re.findall(r"\bPO[-:#]?\s*(\d+)\b", pdf_text)
        ]

        if not po_numbers:
            print("❌ No PO numbers found in file")
            return

        print(f"📄 Found POs: {po_numbers}")
        po_k_map = extract_po_k_mapping(pdf_text)

        if not po_k_map:
            print("❌ Could not map Packing Lists to POs")
            return

        print(f"🔗 PO → Packing List mapping: {po_k_map}")

        po_text_map = split_pdf_by_po(pdf_text, list(po_k_map.keys()))
        print(po_text_map)
        # --- Step 5: Open Excel ONCE ---
        url_download = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{HACH_FILE}/content"
        resp = HTTP.get(url_download, headers=headers, timeout=60)
        resp.raise_for_status()
        wb = load_workbook(io.BytesIO(resp.content))

        total_updated = 0

        # --- Step 6: Process each PO ---
        for po_number_digits, po_text in po_text_map.items():
            print(f"➡️ Processing PO {po_number_digits}")
            po_k_number = po_k_map.get(po_number_digits)

            if not po_k_number:
                print(f"⚠️ No Packing List mapped for PO {po_number_digits}")
                continue

            print(f"🔗 PO {po_number_digits} → Packing List {po_k_number}")

            ws = wb[po_number_digits]

            tables = list(ws.tables.values())

            table = tables[0]
            start_cell, end_cell = table.ref.split(":")
            start_row = ws[start_cell].row
            start_col = ws[start_cell].column
            end_row = ws[end_cell].row
            end_col = ws[end_cell].column

            data = [
                list(r) for r in ws.iter_rows(
                    min_row=start_row,
                    max_row=end_row,
                    min_col=start_col,
                    max_col=end_col,
                    values_only=True
                )
            ]

            df = pd.DataFrame(data[1:], columns=data[0])
            df["Code"] = df["Code"].astype(str).str.strip()

            code_quantity_map = {}

            for code in df["Code"]:
                code_str = str(code).strip()
                pattern = re.compile(
                    rf"{re.escape(code_str)}\s+(\d+(?:[.,]\d+)?)"
                )

                match = pattern.search(po_text)
                if match:
                    qty_str = match.group(1).replace(",", ".")
                    try:
                        quantity = float(qty_str)
                    except ValueError:
                        quantity = None
                    code_quantity_map[code_str] = quantity
                else:
                    code_quantity_map[code_str] = None
            updated = 0

            for idx, row in df.iterrows():
                code = str(row["Code"]).strip()
                if code not in po_text:
                    continue

                if is_empty(row.get("Packing List")):
                    df.at[idx, "Packing List"] = po_k_number

                if is_empty(row.get("რა რიცხვში გამოგზავნეს Packing List-ი")):
                    df.at[idx, "რა რიცხვში გამოგზავნეს Packing List-ი"] = confirmation_date_str

                if is_empty(row.get("ჩამოსვლის სავარაუდო თარიღი")):
                    df.at[idx, "ჩამოსვლის სავარაუდო თარიღი"] = arrival_date_str

                if is_empty(row.get("რამდენი გამოიგზავნა")):
                    df.at[idx, "რამდენი გამოიგზავნა"] = code_quantity_map.get(code)

                updated += 1

            if updated == 0:
                print(f"⚠️ No matching codes for PO {po_number_digits}")
                continue

            total_updated += updated
            mask_coo = df["Code"] == "CoO"
            df.loc[mask_coo, "Packing List"] = df.loc[mask_coo, "Packing List"].bfill()
            df.loc[mask_coo, "რა რიცხვში გამოგზავნეს Packing List-ი"] = df.loc[mask_coo, "რა რიცხვში გამოგზავნეს Packing List-ი"].bfill()
            df.loc[mask_coo, "ჩამოსვლის სავარაუდო თარიღი"] = df.loc[mask_coo, "ჩამოსვლის სავარაუდო თარიღი"].bfill()
            df.loc[mask_coo, "რამდენი გამოიგზავნა"] = df.loc[mask_coo, "რამდენი გამოიგზავნა"].bfill()
            for r_idx, row in enumerate(df.values.tolist(), start=start_row + 1):
                for c_idx, value in enumerate(row, start=start_col):
                    ws.cell(row=r_idx, column=c_idx).value = value

            print(f"✅ PO {po_number_digits}: {updated} rows updated")

        if total_updated == 0:
            print("⚠️ No updates made to Excel")
            return

        # --- Step 7: Upload Excel ONCE ---
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        upload_url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{HACH_FILE}/content"

        for attempt in range(8):
            resp = HTTP.put(upload_url, headers=headers, data=output.getvalue())
            if resp.status_code in (409, 423):
                time.sleep(min(30, 2 ** attempt))
                continue
            resp.raise_for_status()
            print(f"🎉 Packing List updated successfully ({total_updated} rows)")
            return
    conn.close()

def process_khrone_packing_list(mailbox, message_id, message_date, internet_id):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    try:
        cursor.execute(
            "INSERT INTO processed_messages (internet_id) VALUES (?)",
            (internet_id,)
        )
        conn.commit()
    except sqlite3.IntegrityError:
        print("⚠️ Duplicate email skipped")
        conn.close()
        return
    print(f"Mailbox: {mailbox}")
    print(f"message_id: {message_id}")
    print(f"message_date: {message_date}")

    if isinstance(message_date, str):
        dt = datetime.fromisoformat(message_date.replace("Z", "+00:00"))
    elif isinstance(message_date, datetime):
        dt = message_date
    else:
        print(f"⚠️ Unexpected message_date type: {type(message_date)}")
        return

    confirmation_date = dt.date()

    with EXCEL_LOCK:
        # --- Step 1: Download current orders Excel file ---
        file_stream = None
        wb = None
        orders_df = pd.DataFrame()

        url_download = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{FILE_ID}/content"
        headers = {"Authorization": f"Bearer {ACCESS_TOKEN_DRIVE or One_Drive_Auth()}"}

        max_attempts = 6
        for attempt in range(max_attempts):
            try:
                resp = HTTP.get(url_download, headers=headers, timeout=60)
                resp.raise_for_status()
                file_stream = io.BytesIO(resp.content)
                wb = load_workbook(file_stream)

                if "მიმდინარე " in wb.sheetnames:
                    ws = wb["მიმდინარე "]
                    orders_df = pd.DataFrame(ws.values)
                    orders_df.columns = orders_df.iloc[0]  # first row as header
                    orders_df = orders_df[1:]              # drop header row
                else:
                    print("⚠️ Worksheet 'მიმდინარე ' not found in orders file.")
                    orders_df = pd.DataFrame()
                break
            except Exception as e:
                wait = min(5 * (attempt + 1), 30)
                print(f"⚠️ Error downloading main file (attempt {attempt+1}/{max_attempts}): {e}. Sleeping {wait}s")
                time.sleep(wait)
        else:
            print("❌ Gave up downloading orders file after multiple attempts")
            return

        # --- Step 2: Get email attachments ---
        att_url = f"https://graph.microsoft.com/v1.0/users/{mailbox}/messages/{message_id}/attachments"
        att_resp = HTTP.get(att_url, headers=get_headers(), timeout=20)
        if att_resp.status_code != 200:
            print(f"❌ Error fetching attachments: {att_resp.status_code} - {att_resp.text}")
            return

        attachments = att_resp.json().get("value", [])

        # --- Step 3: Collect ALL Khrone packing list PDFs ---
        packing_pdfs = []

        for att in attachments:
            name = att.get("name", "")
            if (
                name.lower().endswith(".pdf")
                and re.search(r"\d{3}-\d{6}.*(copy\s*packing\s*list|plc)", name, re.IGNORECASE)
                and "contentBytes" in att
            ):
                packing_pdfs.append(att)

        if not packing_pdfs:
            print("⚠️ No Khrone packing list PDFs found")
            return

        # --- Extract PO → items mapping ---
        po_items_map = {}

        def parse_quantity(qty_str):
            try:
                return int(float(qty_str.replace(",", ".")))
            except:
                return None

        for pdf_att in packing_pdfs:

            print(f"📎 Processing packing list: {pdf_att.get('name')}")

            pdf_bytes = base64.b64decode(pdf_att["contentBytes"])
            all_text = ""

            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                for page in pdf.pages:
                    all_text += (page.extract_text() or "") + "\n"

            # --- Extract PO number ---
            po_match = re.search(r"Your Order\s*:?\s*(PO-\d+)", all_text)
            po_number = po_match.group(1) if po_match else None

            if not po_number:
                print("⚠️ PO number not found in this PDF")
                continue
            print(f"🎯 Found PO: {po_number}")

            pdf_lines = [line.strip() for line in all_text.splitlines() if line.strip()]
            items = []
            # Only check codes belonging to this PO
            po_codes = orders_df.loc[orders_df["PO"] == po_number, "Code"]
            for code in po_codes:
                found = False

                for i, line in enumerate(pdf_lines):
                    if re.search(rf"\b{re.escape(str(code))}\b", line):

                        # Look 1–3 lines above
                        for j in range(1, 4):
                            if i - j < 0:
                                continue

                            qty_line = pdf_lines[i - j]
                            qty_match = re.search(r"(\d+,\d+)", qty_line)

                            if qty_match:
                                qty = parse_quantity(qty_match.group(1))
                                if qty is not None:
                                    items.append({"Code": code, "Quantity": qty})
                                    print(f"✅ Matched code {code} with quantity {qty}")
                                    found = True
                                    break
                        if not found:
                            print(f"⚠️ Quantity not found near code {code}")
                        break
                if not found:
                    print(f"⚠️ Code {code} not found in PDF")
            if items:
                po_items_map[po_number] = items
        updated_rows = 0

        for po_number, items in po_items_map.items():

            print(f"\n🔄 Updating Excel for PO {po_number}")

            for item in items:
                code = item["Code"]
                qty = item["Quantity"]

                mask = (orders_df["PO"] == po_number) & (orders_df["Code"] == code)

                if mask.any():
                    for idx in orders_df.index[mask]:

                        # Update quantity
                        orders_df.at[idx, "რეალურად გამოგზავნილი რაოდენობა"] = qty

                        # Update ready date (overwrite logic same as before)
                        current_val = orders_df.at[idx, "ტვირთი მზადაა ასაღებად"]

                        if pd.isna(current_val) or current_val == "":
                            orders_df.at[idx, "ტვირთი მზადაა ასაღებად"] = confirmation_date

                        updated_rows += 1
                        print(f"✅ Updated code {code} for PO {po_number}")

                else:
                    print(f"⚠️ Code {code} not found in Excel for PO {po_number}")

        if updated_rows == 0:
            print("⚠️ No rows updated from packing lists")

        # --- Step 6: Save back to Excel and upload ---
        ws = wb["მიმდინარე "]

        # Write headers if needed
        for col_idx, col_name in enumerate(orders_df.columns.tolist(), start=1):
            ws.cell(row=1, column=col_idx).value = col_name

        # Write data values
        for row_idx, row in enumerate(orders_df.values.tolist(), start=2):
            for col_idx, col_name in enumerate(orders_df.columns.tolist(), start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # Only write if cell is empty
                if cell.value in (None, ""):
                    cell.value = row[col_idx - 1]
                    # Format date column if needed
                    if col_name == "ტვირთი მზადაა ასაღებად" and cell.value:
                        cell.number_format = "DD/MM/YYYY"

        # Save workbook to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Upload back
        url_upload = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{FILE_ID}/content"
        max_attempts = 10
        for attempt in range(max_attempts):
            resp = HTTP.put(url_upload, headers=headers, data=output.getvalue())
            if resp.status_code in (423, 409):  # Locked
                wait_time = min(30, 2**attempt) + random.uniform(0, 2)
                print(f"⚠️ File locked (attempt {attempt+1}/{max_attempts}), retrying in {wait_time:.1f}s...")
                time.sleep(wait_time)
                continue
            resp.raise_for_status()
            print(f"✅ Excel updated successfully with Khrone packing list")
            break
    cursor.execute(
        "INSERT INTO processed_messages VALUES (?)",
        (internet_id,)
    )
    conn.close()

def delivery_date_nonhach(salesorder_number: str, skus: list[str], delivery_start: str, delivery_end: str) -> None:
    with EXCEL_LOCK:
        file_stream = None
        wb = None

        try:
            # --- Step 1: Download Excel ---
            url_download = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{FILE_ID}/content"
            headers = {"Authorization": f"Bearer {ACCESS_TOKEN_DRIVE or One_Drive_Auth()}"}

            for attempt in range(6):
                try:
                    resp = HTTP.get(url_download, headers=headers, timeout=60)
                    resp.raise_for_status()
                    file_stream = io.BytesIO(resp.content)
                    wb = load_workbook(file_stream)
                    break
                except Exception as e:
                    wait = min(5 * (attempt + 1), 30)
                    print(f"⚠️ Download failed ({attempt+1}/6): {e}, retrying in {wait}s")
                    time.sleep(wait)
            else:
                print("❌ Failed to download Excel")
                return

            # --- Step 2: Locate target sheet ---
            target_sheet = None
            target_df = None

            for sheet_name in ("მიმდინარე ", "ჩამოსული"):
                if sheet_name not in wb.sheetnames:
                    continue

                ws = wb[sheet_name]
                df = pd.DataFrame(ws.values)
                df.columns = df.iloc[0]
                df = df[1:]

                if (df["SO"] == salesorder_number).any():
                    target_sheet = sheet_name
                    target_df = df.copy()
                    print(f"📄 Using sheet '{sheet_name}'")
                    break

            if target_sheet is None:
                print(f"⚠️ SO {salesorder_number} not found in any sheet")
                return

            ws = wb[target_sheet]

            # --- Step 3: Validate columns ---
            required_cols = {
                "SO",
                "Code",
                "Supplier Company",
                "შეკვეთის ჩაბარების ვადა"
            }

            if not required_cols.issubset(target_df.columns):
                raise ValueError(f"Missing required columns in '{target_sheet}'")

            # --- Step 4: Normalize ---
            target_df["Code"] = target_df["Code"].astype(str).str.strip()
            target_df["Supplier Company"] = target_df["Supplier Company"].astype(str)
            target_df["შეკვეთის ჩაბარების ვადა"] = (target_df["შეკვეთის ჩაბარების ვადა"].astype(str).str.strip().replace("nan", ""))
            # --- Step 5: Apply delivery dates (SO + SKU, NON-HACH only) ---
            so_sku_mask = (
                (target_df["SO"] == salesorder_number) &
                (target_df["Code"].isin(skus))
            )

            if delivery_start == delivery_end:
                delivery_value = delivery_start
            else:
                delivery_value = f"{delivery_start} – {delivery_end}"
            empty_mask = target_df["შეკვეთის ჩაბარების ვადა"] == ""
            final_mask = so_sku_mask & empty_mask
            target_df.loc[final_mask, "შეკვეთის ჩაბარების ვადა"] = delivery_value
            # --- Step 6: Write back to Excel ---
            for col_idx, col_name in enumerate(target_df.columns, start=1):
                ws.cell(row=1, column=col_idx).value = col_name

            for row_idx, row in enumerate(target_df.itertuples(index=False), start=2):
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col_idx).value = value

            # --- Step 7: Save & upload ---
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            url_upload = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{FILE_ID}/content"

            for attempt in range(10):
                resp = HTTP.put(url_upload, headers=headers, data=output.getvalue())
                if resp.status_code in (423, 409):
                    wait = min(30, 2 ** attempt)
                    print(f"⚠️ File locked, retrying in {wait}s")
                    time.sleep(wait)
                    continue

                resp.raise_for_status()
                print("✅ Excel upload successful")
                return

            raise RuntimeError("Upload failed after retries")

        except Exception as e:
            print(f"❌ Fatal error: {e}")

        finally:
            if wb:
                wb.close()
            if file_stream:
                file_stream.close()
            gc.collect()

def delivery_date_hach(salesorder_number: str,delivery_start: str,delivery_end: str, skus: list[str]) -> None:
    skus = {s.strip().upper() for s in skus}  # normalize once
    delivery_range = (
        delivery_start
        if delivery_start == delivery_end
        else f"{delivery_start} - {delivery_end}"
    )

    with EXCEL_LOCK:
        file_stream = None
        wb = None

        try:
            # --- Step 1: Download HACH Excel ---
            url_download = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{HACH_FILE}/content"
            headers = {
                "Authorization": f"Bearer {ACCESS_TOKEN_DRIVE or One_Drive_Auth()}"
            }

            for attempt in range(6):
                try:
                    resp = HTTP.get(url_download, headers=headers, timeout=60)
                    resp.raise_for_status()
                    file_stream = io.BytesIO(resp.content)
                    wb = load_workbook(file_stream)
                    break
                except Exception as e:
                    wait = min(5 * (attempt + 1), 30)
                    print(f"⚠️ HACH Excel download failed ({attempt+1}/6): {e}, retrying in {wait}s")
                    time.sleep(wait)
            else:
                print("❌ Failed to download HACH Excel")
                return

            # --- Step 2: Find matching sheet by SO number in D4 ---
            target_ws = None

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                cell_value = ws["D4"].value

                if not cell_value:
                    continue

                # Normalize text
                text = str(cell_value).upper()

                # Extract all SO numbers like SO-12345
                found_sos = re.findall(r"SO[-]?\d+", text)

                if salesorder_number.upper() in found_sos:
                    target_ws = ws
                    print(f"📄 HACH sheet matched: '{sheet_name}' (SO found in D4)")
                    break

            if not target_ws:
                print(f"⚠️ SO {salesorder_number} not found in any HACH sheet (checked D4)")
                return


            # --- Step 4: Write delivery date PER SKU (Code in D, Delivery in I) ---
            start_row = 9          # data starts under headers
            code_col_idx = 4       # D
            delivery_col_idx = 9   # I

            for row_idx in range(start_row, target_ws.max_row + 1):
                code_cell = target_ws.cell(row=row_idx, column=code_col_idx).value

                if not code_cell:
                    continue

                code = str(code_cell).strip().upper()
                delivery_cell = target_ws.cell(row=row_idx, column=delivery_col_idx)
                if not delivery_cell.value and (code in skus or code == "COO"):
                    delivery_cell.value = delivery_range

            # --- Step 5: Save & upload ---
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            url_upload = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{HACH_FILE}/content"

            for attempt in range(10):
                resp = HTTP.put(url_upload, headers=headers, data=output.getvalue())

                if resp.status_code in (423, 409):
                    wait = min(30, 2 ** attempt)
                    print(f"⚠️ HACH file locked, retrying in {wait}s")
                    time.sleep(wait)
                    continue

                resp.raise_for_status()
                print("✅ HACH Excel updated successfully")
                return

            raise RuntimeError("❌ HACH Excel upload failed after retries")

        except Exception as e:
            print(f"❌ Fatal error in delivery_date_hach: {e}")

        finally:
            if wb:
                wb.close()
            if file_stream:
                file_stream.close()
            gc.collect()



def send_email(customer_name:str, customer_mail:str, attachments):
    EMAIL_REGEX = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
    if not customer_mail or not isinstance(customer_mail, str):
        print("⚠️ Missing customer email — skipping")
        return
    customer_mail = customer_mail.strip()
    if not EMAIL_REGEX.match(customer_mail):
        print(f"⚠️ Invalid email format: {customer_mail} — skipping")
        return
    #Customers who receive SPECIAL text
    specials = {
        "NEA","UWSCG", "Gardabani TPP", "Gardabani TPP 1",
        "Gardabani TPP2","Georgian Technical University (GTU)","Batumi Water"
    }

    is_special = customer_name in specials
    today_str = date.today().strftime("%d-%m-%Y") 
    # ===== EMAIL CONTENT =====
    if is_special:
        print("SLAAYYY this is special")
        subject = f"შეკვეთა დასრულებულია"
        body = f"""
        <p>მოგესალმებით,</p>
        <p>გატყობინებთ, რომ {today_str}-ში მოხდა თქვენი შეკვეთის მოწოდება. ინფორმაცია მოწოდებული პროდუქციის შესახებ მოცემულია მიმაგრებულ ფაილში.</p>
        <p>გთხოვთ, გამოგზავნოთ მიღება-ჩაბარების აქტი ხელმოსაწერად.</p>
        <p>პატივისცემით,<br>შპს „საქართველოს წყლის სისტემები“, 405310088.</p>
        """
    else:
        print("NOT slay, not special")
        subject = f"შეკვეთა დასრულებულია"
        body = f"""
        <p>მოგესალმებით,</p>
        <p>გატყობინებთ, რომ {today_str}-ში მოხდა თქვენი შეკვეთის მოწოდება. ინფორმაცია მოწოდებული პროდუქციის შესახებ მოცემულია მიმაგრებულ ფაილში.</p>
        <p>გთხოვთ, უზრუნველყოთ ანგარიშსწორება შეთანხმების მიხედვით.</p>
        <p>პატივისცემით,<br>შპს „საქართველოს წყლის სისტემები“, 405310088.</p>
        """
    for from_email in MAILBOXES_2:
        r = HTTP.post(
            f"https://graph.microsoft.com/v1.0/users/{from_email}/sendMail",
            headers={
                "Authorization": f"Bearer {ACCESS_TOKEN_DRIVE}",
                "Content-Type": "application/json"
            },
            json={
            "message": {
                "subject": subject,
                "body": {"contentType": "HTML", "content": body},
                "toRecipients": [{"emailAddress": {"address": customer_mail}}],
                "attachments": attachments
            },
            "saveToSentItems": True
        }
        )
        r.raise_for_status()  # will raise if sending fails
# ==========ENDPOINTS========
@app.route("/")
def index():
    return "App is running. Scheduler is active."
@app.route("/purchase", methods=["POST"])
def purchase_webhook():
    try:
        One_Drive_Auth()

        if not verify_zoho_signature(request, "purchaseorders"):
            return "Invalid signature", 403

        order_id = request.json.get("data", {}).get("purchaseorders_id")
        try:
            append_dataframe_to_table(get_purchase_order_df(order_id), "მიმდინარე ")
            return "OK", 200
        except Exception as e:
            return f"Processing error: {e}", 500


    except Exception as e:
        print(f"❌ Webhook processing error: {e}")
        traceback.print_exc()
        return f"Processing error: {e}", 500
@app.route("/receive", methods=["POST"])
def receive_webhook():
    try:
        One_Drive_Auth()
        if not verify_zoho_signature(request, "purchasereceive"):
            print("❌ Signature verification failed")
            return "Invalid signature", 403
        payload = request.json or {}
        data = payload.get("data", {})
        receive_id = data.get("purchase_receive_id")
        url = f"https://www.zohoapis.com/inventory/v1/purchasereceives/{receive_id}"
        headers = {
        "Authorization": f"Zoho-oauthtoken {ACCESS_TOKEN or refresh_access_token()}"
        }
        response = HTTP.get(url, headers=headers)
        response.raise_for_status()

        receive = response.json().get("purchasereceive", {})
        # --- Extract line items ---
        items = receive.get("line_items", [])
        if not items:
            print("⚠️ No line items found")
        vendor_name = receive.get("vendor_name").upper()
        vendor_name = receive.get("vendor_name", "").upper()
        if vendor_name == "HACH":
            print("🏭 HACH vendor detected")
            POOL.submit(recieved_hach, receive.get("purchaseorder_number"), receive.get("date"),receive.get("line_items", []))
        else:
            POOL.submit(recieved_nonhach, receive.get("purchaseorder_number"), receive.get("date"), receive.get("line_items", []))
        return "OK", 200

    except Exception as e:
        print(f"❌ Webhook processing error: {e}")
        traceback.print_exc()
        return f"Processing error: {e}", 500
@app.route('/delivered', methods=['POST'])
def delivered_webhook():
    One_Drive_Auth()
    if not verify_zoho_signature(request, "shipmentorders"):
            return "Invalid signature", 403
    order_num = request.json.get("data", {}).get("sales_order_number")
    package_num = request.json.get("data", {}).get("package_number")
    package_id = request.json.get("data", {}).get("package_id")
    customer_name = request.json.get("data", {}).get("customer_name")
    customer_mail = request.json.get("data", {}).get("customer_mail")
    print(order_num)
    print(customer_name)
    print(customer_mail)
    headers = {
        "Authorization": f"Zoho-oauthtoken {ACCESS_TOKEN or refresh_access_token()}"
    }

    r = HTTP.get(
        f"https://www.zohoapis.com/inventory/v1/packages/{package_id}",
        headers=headers,
        params={
            "organization_id": ORG_ID,
            "accept": "pdf"
        }
    )

    r.raise_for_status()
    pdf_bytes = r.content

    attachments = []
    if package_id:
        pdf_bytes = r.content
        pdf_base64 = base64.b64encode(pdf_bytes).decode("utf-8")
        attachments.append({
            "@odata.type": "#microsoft.graph.fileAttachment",
            "name": f"{package_num}.pdf",
            "contentType": "application/pdf",
            "contentBytes": pdf_base64
        })
    packages_resp = HTTP.get(
    "https://www.zohoapis.com/inventory/v1/packages",
    headers=headers,
    params={
        "organization_id": ORG_ID,
        "salesorder_number_contains": order_num,  # filter by SO number
        "status": "delivered"
    }
    )
    packages_resp.raise_for_status()
    all_packages = packages_resp.json().get("packages", [])
    matching_packages = [
        pkg for pkg in all_packages
        if str(pkg.get("salesorder_number", "")).strip() == str(order_num).strip()
    ]
    aggregated_items = defaultdict(float)
    for pkg in matching_packages:
        pkg_id = pkg.get("package_id")
        
        if not pkg_id:
            continue

        pkg_resp = HTTP.get(
            f"https://www.zohoapis.com/inventory/v1/packages/{pkg_id}",
            headers=headers,
            params={"organization_id": ORG_ID}
        )
        pkg_resp.raise_for_status()

        pkg_data = pkg_resp.json().get("package", {})
        for item in pkg_data.get("line_items", []):
            sku = item.get("sku", "").strip().upper()
            qty = float(item.get("quantity", 0))
            aggregated_items[sku] += qty
    # Convert to list for process_shipment
    items = [{"sku": sku, "quantity": qty} for sku, qty in aggregated_items.items()]

    email_future = POOL.submit(send_email, customer_name, customer_mail, attachments)
    process_future = POOL.submit(process_shipment, order_num, items)
    try:
        # Get results or raise exceptions
        email_result = email_future.result(timeout=30)
        process_result = process_future.result(timeout=30)
        return "OK", 200
    except Exception as e:
        # Log which task failed
        if email_future.exception():
            print(f"Email sending failed: {email_future.exception()}")
        if process_future.exception():
            print(f"Process shipment failed: {process_future.exception()}")
        return f"Processing error: {e}", 500
@app.route("/invoice", methods=["POST"])
def invoice_webhook():
    One_Drive_Auth()

    if not verify_zoho_signature(request, "invoice"):
        print("❌ Signature verification failed")
        return "Invalid signature", 403

    payload = request.get_json(force=True)
    data = payload.get("data", {})
    so_number = data.get("so_number")

    base_datetime = datetime.now()

    headers = {
        "Authorization": f"Zoho-oauthtoken {ACCESS_TOKEN or refresh_access_token()}",
        "X-com-zoho-inventory-organizationid": ORG_ID
    }

    # 1️⃣ Find Sales Order ID
    search_resp = HTTP.get(
        "https://www.zohoapis.com/inventory/v1/salesorders",
        headers=headers,
        params={"salesorder_number": so_number}
    )
    search_resp.raise_for_status()

    salesorders = search_resp.json().get("salesorders", [])
    so_id = next(
        (so["salesorder_id"] for so in salesorders
         if so.get("salesorder_number") == so_number),
        None
    )

    if not so_id:
        return jsonify({"error": f"Sales Order {so_number} not found"}), 404

    # 2️⃣ Fetch full Sales Order
    so_resp = HTTP.get(
        f"https://www.zohoapis.com/inventory/v1/salesorders/{so_id}",
        headers=headers
    )
    so_resp.raise_for_status()
    so_detail = so_resp.json().get("salesorder", {})

    # 3️⃣ Read delivery lead time CF
    delivery_cf = (
        so_detail
        .get("custom_field_hash", {})
        .get("cf_delivery_after_payment", "")
    )

    if not delivery_cf:
        return jsonify({
            "ok": True,
            "message": "No delivery lead time defined"
        }), 200

    # 4️⃣ Parse weeks (single or range)
    match = re.search(r"(\d+)(?:\s*-\s*(\d+))?\s*(weeks?|კვირ\w*)", delivery_cf.lower())

    if not match:
        return jsonify({
            "ok": True,
            "message": "Delivery lead time format not recognized"
        }), 200

    start_w = int(match.group(1))
    end_w = int(match.group(2)) if match.group(2) else start_w

    start_date = base_datetime + timedelta(weeks=start_w)
    end_date = base_datetime + timedelta(weeks=end_w)

    start_str = start_date.strftime("%d/%m/%Y")
    end_str = end_date.strftime("%d/%m/%Y")

    # 5️⃣ Split items by HACH / NON-HACH (EXCEL-BASED)
    hach_skus = []
    non_hach_skus = []

    hach_reference = load_hach_reference_values()  # Excel first column → SET

    for item in so_detail.get("line_items", []):
        sku = item.get("sku")
        code = item.get("custom_field_hash", {}).get("cf_code")

        if not sku or not code:
            continue

        normalized_code = str(code).strip().upper()

        if normalized_code in hach_reference:
            hach_skus.append(sku.upper())
        else:
            non_hach_skus.append(sku.upper())

    # 6️⃣ Update NON-HACH (SO + SKU)
    if non_hach_skus:
        POOL.submit(delivery_date_nonhach, so_number, non_hach_skus, start_str, end_str)

    # 7️⃣ Update HACH (sheet discovery by SO inside sheet)
    if hach_skus:
        POOL.submit(delivery_date_hach,  so_number, start_str, end_str, hach_skus)

    return "OK", 200

# ===========MAIL PROCESSING============
def safe_request(method, url, **kwargs):
    timeout = kwargs.pop("timeout", 30)
    func = getattr(HTTP, method.lower())
    return func(url, timeout=timeout, **kwargs)
def clear_all_subscriptions():
    headers = get_headers()
    subs_url = f"{GRAPH_URL}/subscriptions"
    resp = safe_request("get", subs_url, headers=headers)
    if resp.status_code != 200:
        raise RuntimeError(f"Failed to list subscriptions: {resp.text}")

    subs = resp.json().get("value", [])
    print(f"Found {len(subs)} existing subscriptions")
    for sub in subs:
        sub_id = sub["id"]
        del_url = f"{GRAPH_URL}/subscriptions/{sub_id}"
        dresp = safe_request("delete", del_url, headers=headers)
        if dresp.status_code not in (202, 204):
            print(f"Could not delete {sub_id}: {dresp.text}")
        else:
            print(f"Deleted subscription {sub_id}")
def create_subscription_for_user(mailbox):
    expiration_time = (datetime.utcnow() + timedelta(minutes=4230)).isoformat() + "Z"
    data = {
        "changeType": "created",
        "notificationUrl": WEBHOOK_URL,
        "resource": f"users/{mailbox}/messages",
        "expirationDateTime": expiration_time,
        "clientState": "secretClientValue"
    }

    print(f"Creating subscription for {mailbox}...")
    try:
        response = safe_request("post", f"{GRAPH_URL}/subscriptions", headers=get_headers(), json=data, timeout=30)
    except Exception as e:
        print(f"❌ Network error creating subscription for {mailbox}: {e}")
        return None

    if response.status_code in (200, 201):
        sub_info = response.json()
        print(f"✅ Created subscription for {mailbox}: {sub_info.get('id')}")
        return sub_info
    elif response.status_code == 202:
        # Accepted. Graph may be validating the endpoint. Return whatever Graph sent.
        print(f"⏳ Subscription for {mailbox} accepted (202). Graph is validating the notification URL.")
        try:
            return response.json()
        except Exception:
            return {}
    else:
        print(f"❌ Failed to create subscription for {mailbox}: {response.status_code} {response.text}")
        return None
def initialize_subscriptions():
    print("[initialize_subscriptions] Setting up subscriptions...")
    clear_all_subscriptions()
    futures = []
    for mailbox in MAILBOXES:
        futures.append(POOL.submit(create_subscription_for_user, mailbox))

    successful_subs = []
    for future in as_completed(futures):
        try:
            result = future.result()
            if result:
                # resource might be in format "users/{mailbox}/messages" - handle robustly
                resource = result.get('resource', '')
                parts = resource.split('/')
                mailbox_name = parts[1] if len(parts) > 1 else None
                successful_subs.append((mailbox_name or "unknown", result.get('id')))
        except Exception as e:
            print(f"❌ Error creating subscription: {e}")

    print(f"\n✅ Successfully created {len(successful_subs)}/{len(MAILBOXES)} subscriptions")
    return successful_subs
def with_app_ctx_call(fn, *args, **kwargs):
    with app.app_context():
        return fn(*args, **kwargs)

# ===========MAIL ENDPOINTS============
@app.route("/webhook", methods=["GET", "POST"])
def webhook():
    # --- Validation: Graph sends GET with validationToken param ---
    validation_token = request.args.get("validationToken")
    if validation_token:
        print(f"Validation request received: {validation_token}")
        resp = make_response(validation_token, 200)
        resp.mimetype = "text/plain"
        return resp

    if request.method != "POST":
        return jsonify({"status": "active"}), 200
    
    try:
        data = request.json or {}
        notifications = data.get("value", [])

        # --- Patterns ---
        po_pattern = re.compile(
            r'(?i)(?:purchase order\s+|order confirmation\s+)?'
            r'PO\s*[-:#–]?\s*\d+\b'
            r'(?![^\n]*\bhas been (?:partially\s*)?received\b)'
        )

        greenlight_pattern = re.compile(
            r'(Greenlight|Shipping)\s+request.*?/\s*K\d{6,}',
            re.IGNORECASE
        )

        khrone_oa_pattern = re.compile(
            r'O/A\s+for\s+order\s+PO-\d+',
            re.IGNORECASE
        )

        for notification in notifications:
            resource = notification.get("resource", "")
            message_url = f"{GRAPH_URL}/{resource}?$select=id,internetMessageId,subject,from,toRecipients,ccRecipients,receivedDateTime,body"

            message_response = safe_request(
                "get",
                message_url,
                headers=get_headers(),
                timeout=20
            )

            if message_response.status_code != 200:
                print(
                    f"❌ Error fetching message: "
                    f"{message_response.status_code} - {message_response.text}"
                )
                continue

            message = message_response.json()
            internet_id = message.get("internetMessageId")

            # --- Message fields ---
            subject = message.get("subject", "").strip()

            sender_email = (
                message.get("from", {})
                .get("emailAddress", {})
                .get("address", "")
                .lower()
            )

            message_id = message.get("id")
            message_date = message.get("receivedDateTime")

            if not message_id:
                continue
            if sender_email in MAILBOXES_2:
                print("↩️ Ignoring self-sent email")
                continue

            to_emails = [
                r.get("emailAddress", {}).get("address", "")
                for r in message.get("toRecipients", [])
            ]

            cc_emails = [
                r.get("emailAddress", {}).get("address", "")
                for r in message.get("ccRecipients", [])
            ]

            # --- Parse mailbox from resource ---
            mailbox = "unknown"
            try:
                path_parts = resource.split("/")
                if len(path_parts) >= 2 and path_parts[0].lower() in ("users", "me"):
                    mailbox = path_parts[1]
            except Exception:
                print(f"⚠️ Unexpected resource format: {resource}")

            # --- Log message ---
            print("📨 New message received")
            print(f"   Subject: {subject}")
            print(f"   From: {sender_email}")
            print(f"   To: {', '.join(to_emails) if to_emails else '—'}")
            if cc_emails:
                print(f"   CC: {', '.join(cc_emails)}")
            print("-" * 60)
            # --- Extract message body text ---
            body_content = ""

            try:
                body_content = (
                    message.get("body", {})
                    .get("content", "")
                    .lower()
                )
            except Exception:
                body_content = ""
            is_khrone = sender_email.endswith("@krohne.com")
            is_hach = sender_email.endswith("@hach.com")
            is_atb = sender_email.endswith("@atbwater.com")
            has_po_generic = re.search(r'PO-\d+', subject, re.IGNORECASE)

            # 1️⃣ KHRONE readiness (single check)
            if is_khrone and "notification of readiness of goods:" in subject.lower():
                print("✅ Khrone packing list → process_khrone_packing_list")
                POOL.submit(
                    process_khrone_packing_list,
                    mailbox,
                    message_id,
                    message_date,
                    internet_id
                )

            # 2️⃣ KHRONE O/A
            elif is_khrone and (khrone_oa_pattern.search(subject) or khrone_oa_pattern.search(body_content)):
                print("✅ Khrone O/A → process_khrone_message")
                POOL.submit(
                    process_khrone_message,
                    mailbox,
                    message_id,
                    message_date,
                    internet_id
                )

            # 3️⃣ HACH
            elif is_hach and subject:
                is_greenlight = greenlight_pattern.search(subject)
                has_po_hach = po_pattern.search(subject)

                if is_greenlight:
                    print("✅ Hach Greenlight → packing_list")
                    POOL.submit(
                        packing_list,
                        mailbox,
                        message_id,
                        message_date,
                        internet_id
                    )

                elif has_po_hach:
                    print("✅ Hach PO confirmation → process_hach_message")
                    POOL.submit(
                        process_hach_message,
                        mailbox,
                        message_id,
                        message_date,
                        internet_id
                    )

                else:
                    print("ℹ️ Hach mail ignored (no PO or Greenlight)")

            # 4️⃣ Generic PO
            elif has_po_generic or is_atb:
                print("↪️ Generic PO mail → process_message")
                POOL.submit(
                    process_message,
                    mailbox,
                    message_id,
                    message_date,
                    internet_id
                )

            # 5️⃣ Ignore everything else
            else:
                print("ℹ️ Mail ignored (no PO or relevant pattern)")

        return jsonify({"status": "accepted"}), 202

    except Exception as e:
        print(f"❌ Error processing webhook: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500
def _initialize_subscriptions_worker(flask_app):
    with flask_app.app_context():
        try:
            initialize_subscriptions()
        except Exception as e:
            print(f"❌ initialize_subscriptions_worker exception: {e}")
@app.route("/init", methods=["GET", "POST"])
def init_subscriptions_endpoint():
    try:
        print("🔄 Starting subscription initialization in background...")
        # Submit worker that establishes app context itself.
        POOL.submit(_initialize_subscriptions_worker, app)
        return jsonify({
            "status": "success",
            "message": "Subscription initialization started in background"
        }), 200
    except Exception as e:
        print(f"❌ Initialization failed: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500
@app.route("/subscriptions", methods=["GET"])
def list_subscriptions():
    try:
        resp = safe_request("get", f"{GRAPH_URL}/subscriptions", headers=get_headers(), timeout=20)
        if resp.status_code == 200:
            subs = resp.json().get("value", [])
            return jsonify({
                "status": "success",
                "count": len(subs),
                "subscriptions": subs
            }), 200
        else:
            return jsonify({"status": "error", "message": resp.text}), 400
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
@app.route("/cleanup", methods=["GET", "POST"])
def cleanup_subscriptions():
    try:
        print("🧹 Cleaning up subscriptions...")
        # Run cleanup in background to avoid blocking
        POOL.submit(with_app_ctx_call, clear_all_subscriptions)
        return jsonify({"status": "success", "message": "Subscription cleanup scheduled"}), 200
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

# ========== HEALTH CHECK ===============
@app.route("/health")
def health():
    return {'health':'ok'}
